from __future__ import annotations

import json
import tempfile
import textwrap
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from content_integrity.detectors import built_in_llm_rules, build_tortured_rule_index, detect_llm_trace, detect_tortured_phrases, load_tortured_rules
from content_integrity.models import Finding, ParsedRecord
from content_integrity.pipeline import _aggregate_findings, run_default_pipeline
from content_integrity.reporting import _normalized_doi, _risk_priority_key, build_content_integrity_frontend_json
from content_integrity.template_matching_common import _candidate_pairs, _content_class, _similarity, build_normalized_text, build_skeleton_text
from content_integrity.validators import ContextValidator
from content_integrity.validators.context_validator import _parse_validator_payload
from content_integrity.xml_parser import parse_xml, parse_xml_records
from content_integrity.utils import dedupe_records

ROOT = Path(__file__).resolve().parents[1]


def _write_temp_file(directory: Path, name: str, content: str) -> Path:
    path = directory / name
    path.write_text(textwrap.dedent(content).strip(), encoding="utf-8")
    return path


def _tortured_finding(status: str = "", severity: str = "medium") -> Finding:
    return Finding(
        finding_id="FND-00001",
        record_id="REC-1",
        source_file="sample.xml",
        detector_type="tortured_phrase",
        category="tortured_phrase",
        matched_text="nervous network",
        evidence_snippet="The model uses a nervous network.",
        section_or_field="abstract_text",
        severity=severity,
        confidence=0.87,
        rule_id="TP-00001",
        expected_term="neural network",
        validation_status=status,
    )


def _submission(report: dict, abstract_id: str) -> dict:
    return next(item for item in report.values() if item["abstract_id"] == abstract_id)


def _check(submission: dict, name: str) -> dict:
    return next(item for item in submission["checks"] if item["check_name"] == name)


def _record_supporting_checks(submission: dict) -> dict[str, dict]:
    template = _check(submission, "template_detection")
    record = next(item for item in template["result"]["supporting_data"] if item["evidence_scope"] == "RECORD")
    return {item["check_name"]: item for item in record["supporting_checks"]}


def _summary_data(submission: dict) -> dict:
    return _check(submission, "content_integrity_summary")["result"]["supporting_data"][0]


class PipelineTests(unittest.TestCase):
    def test_integration_doi_key_is_canonical(self) -> None:
        self.assertEqual(_normalized_doi(" HTTPS://doi.org/10.1002/CAM4.2839 "), "10.1002/cam4.2839")

    def test_parser_splits_bundled_asco_sub_articles(self) -> None:
        path = self._temp_xml(
            """
            <article>
              <front><journal-meta><journal-title-group><journal-title>JCO</journal-title></journal-title-group></journal-meta>
                <article-meta>
                  <article-id custom-type="abstract-id">e1</article-id>
                  <title-group><article-title>Root abstract</article-title></title-group>
                  <contrib-group><contrib contrib-type="presenter"><name><surname>Root</surname><given-names>A</given-names></name></contrib><aff>Root Hospital</aff></contrib-group>
                  <copyright-year>2026</copyright-year>
                  <abstract><p><bold>e1</bold></p><p><bold>Background: </bold>Root background. <bold>Methods: </bold>Root methods. NCT12345678.</p></abstract>
                  <kwd-group><kwd>root-keyword</kwd></kwd-group>
                </article-meta>
              </front>
              <sub-article article-type="meeting-abstract">
                <front-stub><journal-meta><journal-title-group><journal-title>JCO</journal-title></journal-title-group></journal-meta>
                  <article-meta>
                    <article-id custom-type="abstract-id">e2</article-id>
                    <article-categories><subj-group><subject>Breast Cancer—Metastatic</subject></subj-group></article-categories>
                    <title-group><article-title>Nested abstract</article-title></title-group>
                    <contrib-group><contrib contrib-type="author"><name><surname>Nested</surname><given-names>B</given-names></name></contrib><aff>Nested Hospital</aff></contrib-group>
                    <copyright-year>2026</copyright-year>
                    <abstract><p><bold>e2</bold></p><p><bold>Results: </bold>Nested results.<table-wrap><table><tr><td>REMOVE TABLE</td></tr></table></table-wrap> <bold>Conclusions: </bold>Nested conclusion.</p></abstract>
                    <kwd-group><kwd>nested-keyword</kwd></kwd-group>
                  </article-meta>
                </front-stub>
              </sub-article>
            </article>
            """
        )
        root, nested = parse_xml_records(path)
        self.assertEqual((root.record_id, nested.record_id), ("e1", "e2"))
        self.assertEqual((root.author_count, nested.author_count), (1, 1))
        self.assertEqual((root.affiliations, nested.affiliations), (["Root Hospital"], ["Nested Hospital"]))
        self.assertEqual((root.keywords, nested.keywords), (["root-keyword"], ["nested-keyword"]))
        self.assertEqual((root.publication_year, nested.publication_year), ("2026", "2026"))
        self.assertFalse(root.abstract_text.startswith("e1"))
        self.assertFalse(nested.abstract_text.startswith("e2"))
        self.assertNotIn("REMOVE TABLE", nested.abstract_text)
        self.assertEqual(root.trial_ids, ["NCT12345678"])
        self.assertTrue(all(block.source_start is not None and block.source_end is not None for block in root.trace_text_blocks))

    def test_template_similarity_is_tokenized_and_symmetric(self) -> None:
        left = "results " + "patient response " * 120
        right = "conclusion " + "patient response " * 120
        self.assertEqual(_similarity(left, right), _similarity(right, left))
        self.assertGreater(_similarity(left, right), 0.99)

    def test_ngram_candidates_recover_reordered_template(self) -> None:
        chunks = [
            "Patients received protocol treatment with prospective clinical assessment and standardized longitudinal outcome collection.",
            "Tumor response and safety endpoints were evaluated by independent reviewers using predefined analysis criteria.",
            "The findings support additional validation in larger diverse populations and future randomized clinical studies.",
        ]
        records = [
            ParsedRecord("a.xml", record_id="A", abstract_text=" ".join(chunks)),
            ParsedRecord("b.xml", record_id="B", abstract_text=" ".join((chunks[1], chunks[0], chunks[2]))),
        ]
        skeletons = {record.record_id: build_skeleton_text(record) for record in records}
        normalized = {record.record_id: build_normalized_text(record) for record in records}
        self.assertIn(("A", "B"), _candidate_pairs(records, skeletons, normalized))

    def test_content_class_preserves_valid_short_abstracts(self) -> None:
        short = ParsedRecord("short.xml", record_id="short", abstract_text="A concise but valid clinical abstract reports treatment outcomes in patients.")
        boilerplate = ParsedRecord("na.xml", record_id="na", abstract_text="N/A N/A N/A")
        self.assertEqual(_content_class(short, build_skeleton_text(short)), "valid_short")
        self.assertEqual(_content_class(boilerplate, build_skeleton_text(boilerplate)), "empty_or_unusable")

    def test_trial_ids_are_masked_before_gene_names(self) -> None:
        record = ParsedRecord("trial.xml", abstract_text="NCT12345678 evaluated BRCA1.")
        self.assertEqual(build_skeleton_text(record), "<TRIAL_ID> evaluated <GENE>.")

    def test_dedupe_keeps_distinct_content_from_same_source(self) -> None:
        records, warnings = dedupe_records([
            ParsedRecord("bundle.xml", record_id="same", abstract_text="First abstract"),
            ParsedRecord("bundle.xml", record_id="same", abstract_text="Different abstract"),
        ])
        self.assertEqual(len(records), 2)
        self.assertEqual(len({record.record_id for record in records}), 2)
        self.assertEqual(warnings[0]["reason"], "record_id_collision_disambiguated")

    def test_parser_handles_article_and_article_set(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir_str:
            temp_dir = Path(temp_dir_str)
            article_path = _write_temp_file(
                temp_dir,
                "article.xml",
                """
                <article article-type="Original Research">
                  <front>
                    <journal-meta>
                      <journal-title-group><journal-title>Test Journal</journal-title></journal-title-group>
                    </journal-meta>
                    <article-meta>
                      <article-id pub-id-type="manuscript">TEST-1</article-id>
                      <article-title>Sample Article</article-title>
                      <abstract>
                        <p><b>Background:</b> Something happened. <b>Methods:</b> We did tests. <b>Results:</b> Great. <b>Conclusion:</b> Works.</p>
                      </abstract>
                      <kwd-group><kwd>alpha</kwd><kwd>beta</kwd></kwd-group>
                      <contrib-group>
                        <contrib contrib-type="presenter"><name><given-names>Primary</given-names><surname>Researcher</surname></name></contrib>
                        <contrib contrib-type="author"><name><given-names>Ada</given-names><surname>Lovelace</surname></name></contrib>
                      </contrib-group>
                      <aff id="aff1"><institution>Test Institute</institution></aff>
                      <history><date date-type="accepted"><year>2026</year></date></history>
                    </article-meta>
                  </front>
                </article>
                """,
            )
            article_set_path = _write_temp_file(
                temp_dir,
                "article_set.xml",
                """
                <article_set dtd_version="4.28.11">
                  <article ms_no="MS-1" tracking_no="T-1">
                    <journal>
                      <full_journal_title>Another Journal</full_journal_title>
                    </journal>
                    <publication_type>Research Article</publication_type>
                    <article_title>Another Sample</article_title>
                    <abstract>
                      <p>Background: First. Methods: Second. Results: Third. Conclusion: Fourth.</p>
                    </abstract>
                    <author_list>
                      <author>
                        <first_name>Grace</first_name><last_name>Hopper</last_name>
                        <affiliation><inst>Example University</inst></affiliation>
                      </author>
                    </author_list>
                    <history><date date-type="accepted"><year>2025</year></date></history>
                  </article>
                </article_set>
                """,
            )
            article_record = parse_xml(article_path)
            article_set_record = parse_xml(article_set_path)
            self.assertEqual(article_record.record_id, "TEST-1")
            self.assertEqual(article_record.journal, "Test Journal")
            self.assertTrue(article_record.structured_abstract)
            self.assertGreaterEqual(article_record.abstract_section_count, 4)
            self.assertEqual(article_record.primary_author, "Primary Researcher")
            self.assertEqual(article_record.authors, ["Primary Researcher", "Ada Lovelace"])
            self.assertEqual(article_set_record.record_id, "MS-1")
            self.assertEqual(article_set_record.journal, "Another Journal")
            self.assertEqual(article_set_record.publication_year, "2025")
            self.assertEqual(article_set_record.primary_author, "Grace Hopper")

    def test_llm_trace_detector_finds_synthetic_trace(self) -> None:
        record = parse_xml(
            self._temp_xml(
                """
                <article article-type="Original Research">
                  <front>
                    <journal-meta><journal-title-group><journal-title>Test</journal-title></journal-title-group></journal-meta>
                    <article-meta>
                      <article-id pub-id-type="manuscript">TEST-2</article-id>
                      <article-title>As an AI language model, this is a test</article-title>
                      <abstract><p>As an AI language model, I cannot provide medical advice.</p></abstract>
                      <history><date date-type="accepted"><year>2026</year></date></history>
                    </article-meta>
                  </front>
                </article>
                """
            )
        )
        findings = detect_llm_trace(record, built_in_llm_rules())
        self.assertTrue(any(finding.rule_id == "LLM-001" for finding in findings))
        self.assertTrue(any(finding.rule_id == "LLM-007" for finding in findings))

    def test_llm_trace_detector_finds_research_backed_weak_residue(self) -> None:
        record = ParsedRecord(
            source_file="sample.xml",
            record_id="TEST-WEAK",
            abstract_text=(
                "It is essential to note that the user requested a revision. "
                "\n### Revised abstract\n---"
            ),
        )

        rule_ids = {finding.rule_id for finding in detect_llm_trace(record, built_in_llm_rules())}

        self.assertTrue({"LLM-025", "LLM-027", "LLM-028"} <= rule_ids)
        self.assertNotIn("LLM-026", rule_ids)

    def test_tortured_phrase_detector_finds_synthetic_phrase(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir_str:
            temp_dir = Path(temp_dir_str)
            dictionary_path = _write_temp_file(
                temp_dir,
                "tortured.csv",
                """
                Fingerprint - Tortured Phrase,Expected Text,Nb Retrieved Papers
                "nervous network","neural network","10"
                "mechanical learning","machine learning","12"
                """,
            )
            rules = load_tortured_rules(dictionary_path)
            record = parse_xml(
                self._temp_xml(
                    """
                    <article article-type="Original Research">
                      <front>
                        <journal-meta><journal-title-group><journal-title>Test</journal-title></journal-title-group></journal-meta>
                        <article-meta>
                          <article-id pub-id-type="manuscript">TEST-3</article-id>
                          <article-title>Mechanical learning is surprising</article-title>
                          <abstract><p>The model uses a nervous network.</p></abstract>
                          <history><date date-type="accepted"><year>2026</year></date></history>
                        </article-meta>
                      </front>
                    </article>
                    """
                )
            )
            findings = detect_tortured_phrases(record, rules, build_tortured_rule_index(rules))
            self.assertTrue(any(finding.matched_text.lower() == "nervous network" for finding in findings))
            self.assertTrue(any(finding.matched_text.lower() == "mechanical learning" for finding in findings))

    def test_tortured_phrase_detector_honors_dictionary_queries_and_sentence_boundaries(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir_str:
            dictionary_path = _write_temp_file(
                Path(temp_dir_str),
                "tortured.csv",
                '''
                Fingerprint - Tortured Phrase,Expected Text,Nb Retrieved Papers
                """surface region"" AND ""surface area""","surface area","1"
                """corridor impact"" AND (""magnetic"" OR ""voltage"")","Hall effect","1"
                """magnetic cell separation"" NOT ""with MACS""","magnetic-activated cell sorting","1"
                """very long tortured phrase""","expected phrase","1"
                """multi token tortured phrase""","expected phrase","1"
                ''',
            )
            rules = load_tortured_rules(dictionary_path)
            record = ParsedRecord(
                source_file="sample.xml",
                record_id="REC-1",
                title="A corridor impact from a voltage sensor",
                abstract_text=(
                    "The surface region was measured without the required context. "
                    "Magnetic cell separation with MACS was used. "
                    "A very long tortured. Phrase must not cross a sentence. "
                    "This multi-token tortured phrase must still be retrieved."
                ),
            )

            findings = detect_tortured_phrases(record, rules, build_tortured_rule_index(rules))

            self.assertEqual(
                [finding.matched_text.lower() for finding in findings],
                ["corridor impact", "multi-token tortured phrase"],
            )
            self.assertTrue(all(rule.rule_id.startswith("TP-") and len(rule.rule_id) == 15 for rule in rules))

    def test_tortured_phrase_proximity_operator_is_enforced_not_stripped(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir_str:
            dictionary_path = _write_temp_file(
                Path(temp_dir_str),
                "tortured.csv",
                '''
                Fingerprint - Tortured Phrase,Expected Text,Nb Retrieved Papers
                """gatherings clustering""~10","clustering","1"
                ''',
            )
            rules = load_tortured_rules(dictionary_path)
            index = build_tortured_rule_index(rules)
            self.assertEqual([rule.proximity for rule in rules], [10])

            def matches(text: str) -> list[str]:
                record = ParsedRecord(source_file="s.xml", record_id="R1", abstract_text=text)
                return [finding.matched_text for finding in detect_tortured_phrases(record, rules, index)]

            self.assertEqual(
                matches("We performed several data gatherings before applying k-means clustering."),
                ["gatherings before applying k-means clustering"],
            )
            self.assertEqual(matches("We used gatherings clustering directly."), ["gatherings clustering"])
            # Outside the ten-token window, and never across a sentence boundary.
            self.assertEqual(matches("gatherings " + "filler " * 12 + "clustering."), [])
            self.assertEqual(matches("We did the gatherings. Then we applied clustering."), [])

    def test_tortured_phrase_proximity_inside_a_not_clause_is_parsed(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir_str:
            dictionary_path = _write_temp_file(
                Path(temp_dir_str),
                "tortured.csv",
                '''
                Fingerprint - Tortured Phrase,Expected Text,Nb Retrieved Papers
                """Chime labs"" NOT ""CHIMe labs Stanford""~5","chime laboratories","1"
                ''',
            )
            rules = load_tortured_rules(dictionary_path)
            index = build_tortured_rule_index(rules)
            self.assertEqual([(item.phrase, item.proximity) for item in rules[0].excluded], [("CHIMe labs Stanford", 5)])

            excluded = ParsedRecord(
                source_file="s.xml",
                record_id="R1",
                abstract_text="Data came from the CHIMe labs Stanford collaboration.",
            )
            flagged = ParsedRecord(
                source_file="s.xml", record_id="R2", abstract_text="The Chime labs reported the result.",
            )
            self.assertEqual(detect_tortured_phrases(excluded, rules, index), [])
            self.assertEqual(
                [finding.matched_text for finding in detect_tortured_phrases(flagged, rules, index)],
                ["Chime labs"],
            )

    def test_tortured_phrase_provenance_and_confidence_basis_are_recorded(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir_str:
            dictionary_path = _write_temp_file(
                Path(temp_dir_str),
                "tortured.csv",
                '''
                Fingerprint - Tortured Phrase,Expected Text,Nb Retrieved Papers
                "nervous network","neural network","10"
                ''',
            )
            rules = load_tortured_rules(dictionary_path, "wiley_tortured_seed_v1")
            row = rules[0].to_dict()

            self.assertEqual(row["source"], "tortured.csv")
            self.assertTrue(row["dictionary_version"].startswith("wiley_tortured_seed_v1+sha256:"))
            self.assertEqual(row["confidence_basis"], "heuristic_rule_strength")
            self.assertEqual(row["rule_strength"], rules[0].severity)
            self.assertEqual(row["expected_term"], "neural network")

            dictionary_path.write_text(
                dictionary_path.read_text(encoding="utf-8") + '"bosom peril","breast cancer","5"\n',
                encoding="utf-8",
            )
            edited = load_tortured_rules(dictionary_path, "wiley_tortured_seed_v1")
            self.assertNotEqual(edited[0].dictionary_version, rules[0].dictionary_version)

    def test_legitimate_scientific_phrasing_is_not_flagged_as_tortured(self) -> None:
        rules = load_tortured_rules(ROOT / "🤷_tortured.csv", "wiley_tortured_seed_v1")
        index = build_tortured_rule_index(rules)
        legitimate = [
            "Patients received pembrolizumab plus chemotherapy for advanced non-small cell lung cancer.",
            "The primary endpoint was progression-free survival assessed by blinded independent central review.",
            "Deep learning models were trained on whole-slide images to predict treatment response.",
            "Hierarchical clustering of gene expression profiles identified three molecular subtypes.",
            "Overall survival was estimated using the Kaplan-Meier method with a two-sided log-rank test.",
            "Neural network architectures were compared against a random forest baseline.",
            "Machine learning classifiers were validated on an independent external cohort.",
            "Circulating tumor DNA was measured at baseline and after two cycles of therapy.",
        ]
        for sentence in legitimate:
            with self.subTest(sentence=sentence):
                record = ParsedRecord(source_file="s.xml", record_id="R1", abstract_text=sentence)
                findings = detect_tortured_phrases(record, rules, index)
                self.assertEqual(
                    [(finding.matched_text, finding.rule_id) for finding in findings], [],
                )

    def test_tortured_phrase_detector_deduplicates_only_exact_matches(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir_str:
            dictionary_path = _write_temp_file(
                Path(temp_dir_str),
                "tortured.csv",
                '''
                Fingerprint - Tortured Phrase,Expected Text,Nb Retrieved Papers
                "nervous network","neural network","1"
                "nervous network","neural network","1"
                """nervous network"" AND ""model""","neural network","1"
                ''',
            )
            rules = load_tortured_rules(dictionary_path)
            record = ParsedRecord(
                source_file="sample.xml",
                record_id="REC-1",
                abstract_text="One nervous network model and another nervous network model.",
            )

            findings = detect_tortured_phrases(record, rules, build_tortured_rule_index(rules))

            self.assertEqual(len(findings), 4)
            self.assertEqual(len({finding.rule_id for finding in findings}), 2)
            self.assertTrue(all(sum(item.rule_id == finding.rule_id for item in findings) == 2 for finding in findings))

    def test_template_pairs_are_reported_for_each_record(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir_str:
            temp_dir = Path(temp_dir_str)
            input_dir = temp_dir / "xmls"
            output_dir = temp_dir / "outputs"
            input_dir.mkdir()
            _write_temp_file(
                input_dir,
                "cluster_a.xml",
                """
                <article article-type="Original Research">
                  <front>
                    <journal-meta><journal-title-group><journal-title>Test Journal</journal-title></journal-title-group></journal-meta>
                    <article-meta>
                      <article-id pub-id-type="manuscript">TPL-A</article-id>
                      <article-title>Template A</article-title>
                      <abstract><p>Background: Patients with advanced disease entered this prospective study. Methods: The primary endpoint was independently assessed response rate. Results: Treatment produced durable responses and improved survival in the enrolled population. Conclusion: These findings support additional controlled investigation.</p></abstract>
                      <history><date date-type="accepted"><year>2026</year></date></history>
                    </article-meta>
                  </front>
                </article>
                """,
            )
            _write_temp_file(
                input_dir,
                "cluster_b.xml",
                """
                <article article-type="Original Research">
                  <front>
                    <journal-meta><journal-title-group><journal-title>Test Journal</journal-title></journal-title-group></journal-meta>
                    <article-meta>
                      <article-id pub-id-type="manuscript">TPL-B</article-id>
                      <article-title>Template B</article-title>
                      <abstract><p>Background: Patients with advanced disease entered this prospective study. Methods: The primary endpoint was independently assessed response rate. Results: Treatment produced durable responses and improved survival in the enrolled population. Conclusion: These findings support additional controlled investigation.</p></abstract>
                      <history><date date-type="accepted"><year>2026</year></date></history>
                    </article-meta>
                  </front>
                </article>
                """,
            )
            dict_path = temp_dir / "dict.csv"
            dict_path.write_text(
                "Fingerprint - Tortured Phrase,Expected Text,Nb Retrieved Papers\n\"nervous network\",\"neural network\",\"1\"\n",
                encoding="utf-8",
            )

            result = run_default_pipeline(input_dir=input_dir, tortured_dictionary_path=dict_path, output_dir=output_dir)

            report = json.loads(result.output_paths["content_integrity_json"].read_text())
            template = _check(_submission(report, "TPL-A"), "template_detection")
            pairs = [item for item in template["result"]["supporting_data"] if item["evidence_scope"] == "PAIR"]
            self.assertEqual(len(pairs), 1)
            pair = pairs[0]
            self.assertEqual(pair["pair_id"], "PAIR-TPL-A--TPL-B")
            self.assertEqual(pair["matched_abstract_id"], "TPL-B")
            self.assertEqual(pair["classification"], "possible_template_reuse")
            self.assertEqual(
                [check["check_name"] for check in pair["sub_checks"]],
                ["exact_text_reuse", "entity_normalized_template", "title_template", "section_similarity"],
            )
            self.assertTrue(next(
                check for check in pair["sub_checks"] if check["check_name"] == "exact_text_reuse"
            )["result"]["supporting_data"])
            self.assertEqual(
                set(_record_supporting_checks(_submission(report, "TPL-A"))),
                {"numerical_contradiction", "design_contradiction", "unverifiable_clinical_trial", "authorship_risk"},
            )

            workbook = load_workbook(result.output_paths["workbook"], read_only=True, data_only=True)
            worksheet = workbook["Check Detail"]
            headers = [cell.value for cell in worksheet[1]]
            details = [dict(zip(headers, row)) for row in worksheet.iter_rows(min_row=2, values_only=True)]
            self.assertEqual({row["Templating (Cross-Author) - Flag"] for row in details}, {"Y"})
            self.assertTrue(all("PAIR-TPL-A--TPL-B" in row["Templating (Cross-Author) - Evidence"] for row in details))
            summaries = {row["record_id"]: row for row in result.abstract_summary_rows}
            self.assertEqual({row["template_flag"] for row in summaries.values()}, {"Yes"})
            self.assertEqual(
                {row["strongest_pair_classification"] for row in summaries.values()},
                {pair["classification"]},
            )
            self.assertEqual(result.abstract_summary_rows[0]["template_cluster_flag"], "No")
            self.assertEqual(result.abstract_summary_rows[1]["template_cluster_flag"], "No")
            metadata = dict(result.run_metadata_rows)
            self.assertEqual(metadata["template_candidate_pair_count"], 1)
            self.assertEqual(metadata["template_final_pair_count"], 1)
            self.assertEqual(metadata["entity_model_inference_count"], 0)

    def test_full_pipeline_creates_only_canonical_outputs_without_deleting_existing_files(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir_str:
            temp_dir = Path(temp_dir_str)
            input_dir = temp_dir / "xmls"
            output_dir = temp_dir / "outputs"
            input_dir.mkdir()
            _write_temp_file(
                input_dir,
                "sample.xml",
                """
                <article article-type="Original Research">
                  <front>
                    <journal-meta><journal-title-group><journal-title>Test Journal</journal-title></journal-title-group></journal-meta>
                    <article-meta>
                      <article-id pub-id-type="manuscript">TEST-6</article-id>
                      <article-title>Simple article</article-title>
                      <abstract><p>Background: Sample. Methods: Sample. Results: Sample. Conclusion: Sample.</p></abstract>
                      <history><date date-type="accepted"><year>2026</year></date></history>
                    </article-meta>
                  </front>
                </article>
                """,
            )
            dict_path = temp_dir / "dict.csv"
            dict_path.write_text(
                "Fingerprint - Tortured Phrase,Expected Text,Nb Retrieved Papers\n\"nervous network\",\"neural network\",\"1\"\n",
                encoding="utf-8",
            )
            output_dir.mkdir()
            legacy_csv = output_dir / "integrity_findings.csv"
            legacy_csv.write_text("user-owned", encoding="utf-8")
            result = run_default_pipeline(input_dir=input_dir, tortured_dictionary_path=dict_path, output_dir=output_dir)
            self.assertTrue(result.output_paths["workbook"].exists())
            self.assertTrue(result.output_paths["content_integrity_json"].exists())
            self.assertEqual(set(result.output_paths), {"content_integrity_json", "workbook"})
            self.assertEqual(result.output_paths["workbook"].name, "Editor_Triage_Workbook.xlsx")
            self.assertEqual(legacy_csv.read_text(encoding="utf-8"), "user-owned")
            self.assertEqual(
                {path.name for path in output_dir.iterdir()},
                {"content_integrity_results.json", "Editor_Triage_Workbook.xlsx", "integrity_findings.csv"},
            )
            with result.output_paths["content_integrity_json"].open(encoding="utf-8") as handle:
                data = json.load(handle)
            self.assertEqual(set(data), {"TEST-6"})
            submission = _submission(data, "TEST-6")
            self.assertEqual(
                [check["check_name"] for check in submission["checks"]],
                ["content_integrity_summary", "tortured_phrases", "llm_response_trace", "template_detection"],
            )
            summary = _check(submission, "content_integrity_summary")
            self.assertEqual(summary["result"]["level"], "LOW")
            self.assertEqual(summary["result"]["supporting_data"][0]["overall_content_risk"], "NONE")
            workbook = load_workbook(result.output_paths["workbook"], read_only=True)
            self.assertEqual(
                workbook.sheetnames,
                ["Dashboard", "High Risk Queue", "Moderate Risk Queue", "Low Risk Queue", "All Abstracts", "Check Detail", "How This Works"],
            )

    def test_frontend_json_scoped_checks_are_reported_and_risk_is_high(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir_str:
            temp_dir = Path(temp_dir_str)
            input_dir = temp_dir / "xmls"
            output_dir = temp_dir / "outputs"
            input_dir.mkdir()
            _write_temp_file(
                input_dir,
                "tortured.xml",
                """
                <article article-type="Original Research">
                  <front>
                    <journal-meta><journal-title-group><journal-title>Test Journal</journal-title></journal-title-group></journal-meta>
                    <article-meta>
                      <article-id pub-id-type="manuscript">TP-1</article-id>
                      <article-title>Tortured phrase example</article-title>
                      <abstract><p>The model uses a nervous network.</p></abstract>
                      <history><date date-type="accepted"><year>2026</year></date></history>
                    </article-meta>
                  </front>
                </article>
                """,
            )
            _write_temp_file(
                input_dir,
                "llm.xml",
                """
                <article article-type="Original Research">
                  <front>
                    <journal-meta><journal-title-group><journal-title>Test Journal</journal-title></journal-title-group></journal-meta>
                    <article-meta>
                      <article-id pub-id-type="manuscript">LLM-1</article-id>
                      <article-title>LLM response example</article-title>
                      <abstract><p>As an AI language model, I cannot provide medical advice.</p></abstract>
                      <history><date date-type="accepted"><year>2026</year></date></history>
                    </article-meta>
                  </front>
                </article>
                """,
            )
            _write_temp_file(
                input_dir,
                "clean.xml",
                """
                <article article-type="Original Research">
                  <front>
                    <journal-meta><journal-title-group><journal-title>Test Journal</journal-title></journal-title-group></journal-meta>
                    <article-meta>
                      <article-id pub-id-type="manuscript">CLEAN-1</article-id>
                      <article-title>Clean abstract</article-title>
                      <abstract><p>No issues detected in this clean abstract.</p></abstract>
                      <history><date date-type="accepted"><year>2026</year></date></history>
                    </article-meta>
                  </front>
                </article>
                """,
            )
            dict_path = temp_dir / "dict.csv"
            dict_path.write_text(
                "Fingerprint - Tortured Phrase,Expected Text,Nb Retrieved Papers\n\"nervous network\",\"neural network\",\"1\"\n",
                encoding="utf-8",
            )
            result = run_default_pipeline(input_dir=input_dir, tortured_dictionary_path=dict_path, output_dir=output_dir)
            with result.output_paths["content_integrity_json"].open(encoding="utf-8") as handle:
                data = json.load(handle)

            self.assertEqual(len(data), 3)
            self.assertEqual(sum(_summary_data(item)["overall_content_risk"] == "HIGH" for item in data.values()), 2)
            self.assertEqual(sum(_summary_data(item)["overall_content_risk"] == "NONE" for item in data.values()), 1)

            tortured_abstract = _submission(data, "TP-1")
            self.assertEqual(len(tortured_abstract["checks"]), 4)
            tortured_check = _check(tortured_abstract, "tortured_phrases")
            self.assertEqual(tortured_check["result"]["level"], "HIGH")
            self.assertEqual(
                tortured_check["result"]["comment"],
                "nervous network → neural network",
            )
            self.assertEqual(_summary_data(tortured_abstract)["overall_content_risk"], "HIGH")
            self.assertEqual(_check(tortured_abstract, "llm_response_trace")["result"]["supporting_data"], [])
            self.assertTrue(all(
                not check["result"]["supporting_data"]
                for check in _record_supporting_checks(tortured_abstract).values()
            ))
            self.assertEqual(
                [item for item in _check(tortured_abstract, "template_detection")["result"]["supporting_data"] if item["evidence_scope"] == "PAIR"],
                [],
            )

            llm_abstract = _submission(data, "LLM-1")
            llm_check = _check(llm_abstract, "llm_response_trace")
            self.assertEqual(
                llm_check["result"]["comment"],
                "As an AI language model, I cannot provide medical advice.",
            )
            self.assertEqual(llm_check["result"]["supporting_data"][0]["category"], "ai_self_identification")
            self.assertEqual(llm_check["result"]["supporting_data"][0]["rule_id"], "LLM-001")
            self.assertIn(
                llm_check["result"]["supporting_data"][0]["validation_status"],
                {"not_validated", "not_required", ""},
            )

            clean_abstract = _submission(data, "CLEAN-1")
            self.assertEqual(_summary_data(clean_abstract)["overall_content_risk"], "NONE")
            self.assertEqual(_check(clean_abstract, "tortured_phrases")["result"]["supporting_data"], [])
            self.assertEqual(_check(clean_abstract, "llm_response_trace")["result"]["supporting_data"], [])

            workbook = load_workbook(result.output_paths["workbook"], read_only=True, data_only=True)
            sheet = workbook["All Abstracts"]
            headers = [cell.value for cell in sheet[1]]
            excel_rows = {
                row["Abstract ID"]: row
                for row in (
                    dict(zip(headers, (cell.value for cell in cells)))
                    for cells in sheet.iter_rows(min_row=2)
                )
            }
            for abstract in data.values():
                excel = excel_rows[abstract["abstract_id"]]
                summary = _summary_data(abstract)
                self.assertEqual(summary["overall_content_risk"].title(), excel["Overall Risk"])
                self.assertEqual("Yes" if summary["review_required"] else "No", excel["Review Required"])
                design = _record_supporting_checks(abstract)["design_contradiction"]
                self.assertEqual(bool(design["result"]["supporting_data"]), excel["Design Contradiction"] == "Y")

    def test_related_duplicate_pair_promotes_highest_severity(self) -> None:
        # Regression: a possible_related_duplicate pair (exact reuse that related-study
        # context does not excuse) is only visible via the enriched pair pipeline, and
        # highest_severity/overall_content_risk must reflect it, not just possible_template_reuse.
        records = [ParsedRecord(source_file="sample.xml", record_id="A", title="A title"), ParsedRecord(source_file="sample.xml", record_id="B", title="B title")]
        enriched_abstract_rows = [
            {"record_id": "A", "family_id": "", "family_size": 0},
            {"record_id": "B", "family_id": "", "family_size": 0},
        ]
        enriched_pair_rows = [
            {
                "pair_id": "PAIR-A--B",
                "left_record_id": "A",
                "right_record_id": "B",
                "left_title": "A title",
                "right_title": "B title",
                "review_priority": "High",
                "editorial_score": 0.9,
                "pair_class": "possible_related_duplicate",
                "primary_evidence": "exact_results_section",
                "supporting_evidence": "",
                "direct_evidence": "",
                "strongest_section": "results",
                "masked_body_similarity": 0.0,
                "original_body_similarity": 0.0,
                "strongest_masked_section_similarity": 0.0,
                "likely_substitutions": "",
                "context_interpretation": "likely_companion_analysis",
            }
        ]
        summary = _aggregate_findings(
            records, [], [], [],
            enriched_pair_rows=enriched_pair_rows,
            enriched_abstract_rows=enriched_abstract_rows,
        )
        summaries = {row["record_id"]: row for row in summary}
        self.assertEqual(summaries["A"]["template_review_priority"], "High")
        self.assertEqual(summaries["A"]["highest_severity"], "High")
        self.assertEqual(summaries["A"]["overall_content_risk"], "High")

    def test_risk_priority_key_orders_worst_first(self) -> None:
        rows = [
            {"record_id": "LOW", "highest_severity": "Low", "total_finding_count": 1},
            {"record_id": "HIGH", "highest_severity": "High", "total_finding_count": 1},
            {"record_id": "MED-MANY", "highest_severity": "Medium", "total_finding_count": 5},
            {"record_id": "MED-FEW", "highest_severity": "Medium", "total_finding_count": 1},
            {"record_id": "NONE", "highest_severity": "None", "total_finding_count": 0},
        ]
        ordered = [row["record_id"] for row in sorted(rows, key=_risk_priority_key)]
        self.assertEqual(ordered, ["HIGH", "MED-MANY", "MED-FEW", "LOW", "NONE"])

    def test_authorship_high_risk_check_promotes_overall_risk(self) -> None:
        # A HIGH-level authorship signal (final_json.json) with no detector findings or
        # template pairs must still surface as real risk, since it's the only signal.
        record = ParsedRecord(source_file="a.xml", record_id="A", doi="10.1/a")
        authorship_checks_by_key = {
            "10.1/a": {
                "author_network": {"level": "HIGH", "comment": "Shared network with 3 retracted papers."},
                "submission_volume": {"level": "LOW", "comment": ""},
            }
        }
        summary = _aggregate_findings(
            [record], [], [], [], authorship_checks_by_key=authorship_checks_by_key
        )[0]
        self.assertEqual(summary["authorship_flag"], "Yes")
        self.assertEqual(summary["authorship_review_priority"], "High")
        self.assertEqual(summary["authorship_reasons"], "Author Network")
        self.assertEqual(summary["highest_severity"], "High")
        self.assertEqual(summary["overall_content_risk"], "High")

    def test_authorship_low_only_checks_do_not_flag_risk(self) -> None:
        record = ParsedRecord(source_file="a.xml", record_id="A", doi="10.1/a")
        authorship_checks_by_key = {"10.1/a": {"submission_volume": {"level": "LOW", "comment": ""}}}
        summary = _aggregate_findings(
            [record], [], [], [], authorship_checks_by_key=authorship_checks_by_key
        )[0]
        self.assertEqual(summary["authorship_flag"], "No")
        self.assertEqual(summary["overall_content_risk"], "None")

    def test_template_candidate_without_reviewer_finding_does_not_flag_templating(self) -> None:
        records = [ParsedRecord(source_file="sample.xml", record_id="A", title="A title"), ParsedRecord(source_file="sample.xml", record_id="B", title="B title")]
        findings: list[Finding] = []
        enriched_abstract_rows = [
            {"record_id": "A", "family_id": "", "family_size": 0},
            {"record_id": "B", "family_id": "", "family_size": 0},
        ]
        enriched_pair_rows = [
            {
                "pair_id": "PAIR-A--B",
                "left_record_id": "A",
                "right_record_id": "B",
                "left_title": "A title",
                "right_title": "B title",
                "review_priority": "None",
                "editorial_score": 0.0,
                "pair_class": "possible_template_reuse",
                "primary_evidence": "",
                "supporting_evidence": "",
                "direct_evidence": "",
                "strongest_section": "",
                "masked_body_similarity": 0.0,
                "original_body_similarity": 0.0,
                "strongest_masked_section_similarity": 0.0,
                "likely_substitutions": "",
                "context_interpretation": "",
            }
        ]
        output = build_content_integrity_frontend_json(
            records=records,
            findings=findings,
            enriched_pair_rows=enriched_pair_rows,
            enriched_abstract_rows=enriched_abstract_rows,
            abstract_summary_rows=_aggregate_findings(
                records, findings, [], [],
                enriched_pair_rows=enriched_pair_rows,
                enriched_abstract_rows=enriched_abstract_rows,
            ),
            operational_issues=[],
            generated_at="2026-01-01T00:00:00Z",
            git_revision="deadbeef",
        )
        self.assertEqual(output["summary"]["total_submissions"], 2)
        self.assertEqual(output["summary"]["high_risk"], 0)
        self.assertEqual(output["abstracts"][0]["checks"]["templating"]["flagged"], False)
        self.assertEqual(output["abstracts"][1]["checks"]["templating"]["flagged"], False)

    def test_authoritative_template_findings_contain_enriched_evidence(self) -> None:
        records = [ParsedRecord(source_file="sample.xml", record_id="A", title="A title"), ParsedRecord(source_file="sample.xml", record_id="B", title="B title")]
        findings: list[Finding] = []
        enriched_abstract_rows = [
            {"record_id": "A", "family_id": "TF-001", "family_size": 3},
            {"record_id": "B", "family_id": "TF-001", "family_size": 3},
        ]
        enriched_pair_rows = [
            {
                "pair_id": "PAIR-A--B",
                "left_record_id": "A",
                "right_record_id": "B",
                "left_title": "A title",
                "right_title": "B title",
                "review_priority": "High",
                "editorial_score": 0.92,
                "pair_class": "possible_template_reuse",
                "primary_evidence": "entity_normalized_template",
                "supporting_evidence": "shared_title | shared_results",
                "direct_evidence": "Shared distinctive text...",
                "strongest_section": "results",
                "masked_body_similarity": 0.94,
                "original_body_similarity": 0.72,
                "strongest_masked_section_similarity": 0.96,
                "likely_substitutions": "gene: TP53 -> BRCA1",
                "matching_text_evidence": "A: Shared distinctive text in A. || B: Shared distinctive text in B.",
                "context_interpretation": "Different study entities with highly similar writing scaffold",
            }
        ]
        output = build_content_integrity_frontend_json(
            records=records,
            findings=findings,
            enriched_pair_rows=enriched_pair_rows,
            enriched_abstract_rows=enriched_abstract_rows,
            abstract_summary_rows=_aggregate_findings(
                records, findings, [], [],
                enriched_pair_rows=enriched_pair_rows,
                enriched_abstract_rows=enriched_abstract_rows,
            ),
            operational_issues=[],
            generated_at="2026-01-01T00:00:00Z",
            git_revision="deadbeef",
        )
        a_row = next(item for item in output["abstracts"] if item["abstract_id"] == "A")
        self.assertTrue(a_row["checks"]["templating"]["flagged"])
        self.assertEqual(a_row["checks"]["templating"]["template_family_id"], "TF-001")
        self.assertEqual(a_row["checks"]["templating"]["family_size"], 3)
        self.assertEqual(a_row["checks"]["templating"]["findings"][0]["matched_abstract_id"], "B")
        self.assertEqual(a_row["checks"]["templating"]["findings"][0]["primary_evidence"], "entity_normalized_template")
        self.assertEqual(a_row["checks"]["templating"]["findings"][0]["supporting_evidence"], ["shared_title", "shared_results"])
        self.assertEqual(a_row["checks"]["templating"]["findings"][0]["direct_evidence"], "Shared distinctive text...")
        self.assertEqual(a_row["checks"]["templating"]["findings"][0]["strongest_section"], "results")
        self.assertEqual(a_row["checks"]["templating"]["findings"][0]["masked_body_similarity"], 0.94)
        self.assertEqual(a_row["checks"]["templating"]["findings"][0]["original_body_similarity"], 0.72)
        self.assertEqual(a_row["checks"]["templating"]["findings"][0]["strongest_section_similarity"], 0.96)
        self.assertEqual(a_row["checks"]["templating"]["findings"][0]["likely_substitutions"], ["gene: TP53 -> BRCA1"])
        self.assertEqual(a_row["checks"]["templating"]["findings"][0]["context_interpretation"], "Different study entities with highly similar writing scaffold")
        self.assertEqual(a_row["checks"]["templating"]["evidence"], "A: Shared distinctive text in A. || B: Shared distinctive text in B.")
        self.assertEqual(
            a_row["checks"]["templating"]["reason"],
            "94% shared structure with B once biomedical entities are normalized (72% shared original text), different author group.",
        )
        self.assertEqual(a_row["checks"]["templating"]["findings"][0]["evidence"], a_row["checks"]["templating"]["evidence"])
        self.assertEqual(a_row["checks"]["templating"]["findings"][0]["reason"], a_row["checks"]["templating"]["reason"])
        pair = a_row["checks"]["templating"]["evidence_pairs"][0]
        self.assertEqual(pair["submitted_abstract_id"], "A")
        self.assertEqual(pair["matched_abstract_id"], "B")
        self.assertEqual(pair["submitted_evidence"], "Shared distinctive text in A.")
        self.assertEqual(pair["matched_evidence"], "Shared distinctive text in B.")

    def test_related_work_is_not_reported_as_templating(self) -> None:
        records = [ParsedRecord(source_file="sample.xml", record_id="A", title="A title"), ParsedRecord(source_file="sample.xml", record_id="B", title="B title")]
        pair = {
            "pair_id": "PAIR-A--B",
            "left_record_id": "A",
            "right_record_id": "B",
            "left_title": "A title",
            "right_title": "B title",
            "review_priority": "Low",
            "editorial_score": 0.85,
            "pair_class": "possible_related_work",
        }
        output = build_content_integrity_frontend_json(
            records=records,
            findings=[],
            enriched_pair_rows=[pair],
            enriched_abstract_rows=[
                {"record_id": "A", "family_id": "", "family_size": 0},
                {"record_id": "B", "family_id": "", "family_size": 0},
            ],
            abstract_summary_rows=_aggregate_findings(
                records,
                [],
                [],
                [],
                enriched_pair_rows=[pair],
                enriched_abstract_rows=[
                    {"record_id": "A", "family_id": "", "family_size": 0},
                    {"record_id": "B", "family_id": "", "family_size": 0},
                ],
            ),
            operational_issues=[],
            generated_at="2026-01-01T00:00:00Z",
            git_revision="deadbeef",
        )
        self.assertEqual(output["summary"]["high_risk"], 0)
        self.assertFalse(output["abstracts"][0]["checks"]["templating"]["flagged"])

    def test_reporting_cannot_promote_ineligible_finding(self) -> None:
        record = ParsedRecord(source_file="sample.xml", record_id="REC-1", title="Rejected")
        rejected = _tortured_finding("rejected", "high")
        failed = _tortured_finding("validation_failed", "high")
        failed.finding_id = "FND-00002"
        findings = [rejected, failed]
        template_abstracts = [{"record_id": "REC-1", "family_id": "", "family_size": 0}]
        summary = _aggregate_findings([record], findings, [], [])[0]

        output = build_content_integrity_frontend_json(
            records=[record],
            findings=findings,
            enriched_pair_rows=[],
            enriched_abstract_rows=template_abstracts,
            abstract_summary_rows=[summary],
            operational_issues=[{
                "component": "tortured_phrase",
                "record_id": "REC-1",
                "source_file": "sample.xml",
                "status": "failed",
                "error_type": "validation_failed",
                "message": "Validator unavailable.",
                "retry_count": 0,
                "recoverable": True,
            }],
            generated_at="2026-01-01T00:00:00Z",
            git_revision="deadbeef",
        )

        abstract = output["abstracts"][0]
        self.assertEqual(abstract["overall_content_risk"], "None")
        self.assertFalse(abstract["checks"]["tortured_phrases"]["flagged"])
        self.assertTrue(abstract["checks"]["tortured_phrases"]["operational_failure"])
        self.assertTrue(all(not finding["active"] for finding in abstract["checks"]["tortured_phrases"]["findings"]))
        self.assertEqual(len(output["operational_issues"]), 1)

    def test_detector_failure_is_not_reported_as_clear(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_temp_file(
                input_dir,
                "record.xml",
                """
                <article><front><article-meta>
                  <article-id pub-id-type="manuscript">FAILED-CHECK</article-id>
                  <article-title>Detector failure</article-title>
                  <abstract><p>A clean abstract.</p></abstract>
                </article-meta></front></article>
                """,
            )
            dictionary = _write_temp_file(
                root,
                "dict.csv",
                "Fingerprint - Tortured Phrase,Expected Text,Nb Retrieved Papers\n",
            )

            with patch(
                "content_integrity.pipeline.detect_numerical_contradictions",
                side_effect=RuntimeError("detector unavailable"),
            ):
                result = run_default_pipeline(input_dir, dictionary, root / "output")

            report = json.loads(result.output_paths["content_integrity_json"].read_text())
            abstract = _submission(report, "FAILED-CHECK")
            numerical = _record_supporting_checks(abstract)["numerical_contradiction"]
            self.assertEqual(numerical["result"]["level"], "UNKNOWN")
            self.assertEqual(numerical["result"]["supporting_data"], [])
            self.assertTrue(any(
                issue["component"] == "numerical_contradiction"
                for issue in _summary_data(abstract)["operational_issues"]
            ))
            workbook = load_workbook(result.output_paths["workbook"], read_only=True, data_only=True)
            sheet = workbook["Check Detail"]
            headers = [cell.value for cell in sheet[1]]
            detail = dict(zip(headers, next(sheet.iter_rows(min_row=2, values_only=True))))
            self.assertEqual(detail["Operational Issues - Flag"], "Y")
            self.assertIn("numerical_contradiction", detail["Operational Issues - Evidence"])

    def test_pipeline_integrates_contradiction_and_trial_checks(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            input_dir = root / "xmls"
            input_dir.mkdir()
            _write_temp_file(
                input_dir,
                "checks.xml",
                """
                <article><front><article-meta>
                  <article-id pub-id-type="manuscript">CHECKS-1</article-id>
                  <article-title>Contradictory trial</article-title>
                  <abstract><p>
                    This study was open-label and double-blind.
                    The objective response rate was 135%.
                    The trial was registered as NCT123.
                  </p></abstract>
                </article-meta></front></article>
                """,
            )
            dictionary = _write_temp_file(
                root,
                "dict.csv",
                "Fingerprint - Tortured Phrase,Expected Text,Nb Retrieved Papers\n",
            )

            result = run_default_pipeline(input_dir, dictionary, root / "outputs")

            data = json.loads(result.output_paths["content_integrity_json"].read_text())
            row = _submission(data, "CHECKS-1")
            supporting = _record_supporting_checks(row)
            self.assertTrue(supporting["numerical_contradiction"]["result"]["supporting_data"])
            self.assertTrue(supporting["design_contradiction"]["result"]["supporting_data"])
            self.assertTrue(supporting["unverifiable_clinical_trial"]["result"]["supporting_data"])
            self.assertEqual(
                _check(row, "content_integrity_summary")["result"]["comment"],
                "Numerical Contradiction, Design Contradiction, Unverifiable Clinical Trial",
            )

    def test_validator_marks_bad_json_validation_failed(self) -> None:
        class BrokenClient:
            def complete(self, *, system: str, user: str, max_tokens: int = 150, temperature: float = 0.0) -> str:
                return "not-json"

        validator = ContextValidator(client=BrokenClient())
        result = validator.validate(_tortured_finding())

        self.assertEqual(result.status, "validation_failed")
        self.assertEqual(result.reason, "Validator response could not be parsed.")
        self.assertEqual(result.finding_id, "FND-00001")

    def test_validator_marks_incomplete_payload_validation_failed(self) -> None:
        class BrokenClient:
            def complete(self, *, system: str, user: str, max_tokens: int = 150, temperature: float = 0.0) -> str:
                return '{"status": "confirmed"}'

        result = ContextValidator(client=BrokenClient()).validate(_tortured_finding())

        self.assertEqual(result.status, "validation_failed")
        self.assertEqual(result.reason, "Validator response could not be parsed.")

    def test_validator_distinguishes_infrastructure_failure_from_uncertainty(self) -> None:
        class FailedClient:
            def complete(self, *, system: str, user: str, max_tokens: int = 150, temperature: float = 0.0) -> str:
                raise RuntimeError("service unavailable")

        failed_finding = _tortured_finding()
        failed = ContextValidator(client=FailedClient()).validate(failed_finding)
        failed_finding.validation_status = failed.status
        failed_finding.validation_reason = failed.reason

        class UncertainClient:
            def complete(self, *, system: str, user: str, max_tokens: int = 150, temperature: float = 0.0) -> str:
                return '{"status":"uncertain","reason":"The phrase may be legitimate in this context."}'

        uncertain_finding = _tortured_finding()
        uncertain = ContextValidator(client=UncertainClient()).validate(uncertain_finding)
        uncertain_finding.validation_status = uncertain.status

        self.assertEqual(failed.status, "validation_failed")
        self.assertIn("infrastructure", failed.reason)
        self.assertEqual(uncertain.status, "uncertain")
        self.assertEqual(uncertain.reason, "The phrase may be legitimate in this context.")
        record = ParsedRecord(source_file="sample.xml", record_id="REC-1")
        failed_summary = _aggregate_findings([record], [failed_finding], [], [])[0]
        uncertain_summary = _aggregate_findings([record], [uncertain_finding], [], [])[0]
        self.assertEqual(failed_summary["tortured_validation_failed_count"], 1)
        self.assertEqual(uncertain_summary["tortured_uncertain_count"], 1)
        self.assertEqual(failed_summary["overall_content_risk"], "None")
        self.assertEqual(uncertain_summary["overall_content_risk"], "None")
        self.assertEqual(failed_summary["review_required"], "No")
        self.assertEqual(uncertain_summary["review_required"], "Yes")
        self.assertFalse(failed_finding.active)
        self.assertFalse(uncertain_finding.active)
        self.assertTrue(uncertain_finding.review_candidate)
        self.assertEqual(failed_finding.to_dict()["validation_status"], "validation_failed")

    def test_rejected_tortured_phrase_does_not_affect_risk(self) -> None:
        record = ParsedRecord(source_file="sample.xml", record_id="REC-1")
        finding = _tortured_finding("rejected", "high")

        summary = _aggregate_findings([record], [finding], [], [])[0]

        self.assertEqual(summary["tortured_rejected_count"], 1)
        self.assertEqual(summary["total_finding_count"], 0)
        self.assertEqual(summary["highest_severity"], "None")
        self.assertEqual(summary["overall_content_risk"], "None")
        self.assertEqual(summary["review_required"], "No")
        self.assertEqual(summary["tortured_phrase_count"], 1)

    def test_tortured_validation_status_risk_matrix(self) -> None:
        record = ParsedRecord(source_file="sample.xml", record_id="REC-1")
        cases = {
            "confirmed": ("tortured_confirmed_count", "Medium", 1),
            "": ("tortured_not_validated_count", "Medium", 1),
            "rejected": ("tortured_rejected_count", "None", 0),
            "uncertain": ("tortured_uncertain_count", "None", 0),
            "validation_failed": ("tortured_validation_failed_count", "None", 0),
            "candidate": ("tortured_candidate_count", "None", 0),
        }

        for status, (count_field, risk, risk_count) in cases.items():
            with self.subTest(status=status or "blank"):
                finding = _tortured_finding(status)
                summary = _aggregate_findings([record], [finding], [], [])[0]
                self.assertEqual(summary[count_field], 1)
                self.assertEqual(summary["overall_content_risk"], risk)
                self.assertEqual(summary["total_finding_count"], risk_count)
                self.assertEqual(summary["tortured_phrase_count"], 1)
                self.assertEqual(
                    summary["review_required"],
                    "Yes" if status in {"confirmed", "", "uncertain", "candidate"} else "No",
                )

    def test_validator_parses_wrapped_json(self) -> None:
        parsed = _parse_validator_payload('Result: {"status":"confirmed","reason":"Trace confirmed."} done')
        self.assertEqual(parsed["status"], "confirmed")

    def test_pipeline_validation_flag_populates_finding_metadata(self) -> None:
        class StubClient:
            def __init__(self) -> None:
                self.max_tokens_requested: int | None = None

            def complete(self, *, system: str, user: str, max_tokens: int = 150, temperature: float = 0.0) -> str:
                self.max_tokens_requested = max_tokens
                payload = {"status": "rejected", "reason": "Standard terminology, not a plausible substitution."}
                return f"```json\n{json.dumps(payload)}\n```"

        with tempfile.TemporaryDirectory() as temp_dir_str:
            temp_dir = Path(temp_dir_str)
            input_dir = temp_dir / "xmls"
            output_dir = temp_dir / "outputs"
            input_dir.mkdir()
            _write_temp_file(
                input_dir,
                "sample.xml",
                """
                <article article-type="Original Research">
                  <front>
                    <journal-meta><journal-title-group><journal-title>Test Journal</journal-title></journal-title-group></journal-meta>
                    <article-meta>
                      <article-id pub-id-type="manuscript">TEST-7</article-id>
                      <article-title>Validator example</article-title>
                      <abstract><p>The model uses a nervous network in the background section.</p></abstract>
                      <history><date date-type="accepted"><year>2026</year></date></history>
                    </article-meta>
                  </front>
                </article>
                """,
            )
            dict_path = temp_dir / "dict.csv"
            dict_path.write_text(
                "Fingerprint - Tortured Phrase,Expected Text,Nb Retrieved Papers\n\"nervous network\",\"neural network\",\"1\"\n",
                encoding="utf-8",
            )

            stub_client = StubClient()
            with patch("content_integrity.pipeline.build_gpt_oss_client", return_value=stub_client):
                result = run_default_pipeline(
                    input_dir=input_dir,
                    tortured_dictionary_path=dict_path,
                    output_dir=output_dir,
                    validate_llm=True,
                )

            tortured_findings = [finding for finding in result.findings if finding.detector_type == "tortured_phrase"]
            self.assertTrue(tortured_findings)
            self.assertEqual(stub_client.max_tokens_requested, 2048)
            self.assertEqual(tortured_findings[0].validation_status, "rejected")
            self.assertEqual(
                tortured_findings[0].validated_by,
                "gpt-oss-20b:context_validator_v2",
            )
            summary = result.abstract_summary_rows[0]
            self.assertEqual(summary["tortured_rejected_count"], 1)
            self.assertEqual(summary["total_finding_count"], 0)
            self.assertEqual(summary["highest_severity"], "None")
            self.assertEqual(summary["overall_content_risk"], "None")
            self.assertEqual(summary["review_required"], "No")

            report = json.loads(result.output_paths["content_integrity_json"].read_text())
            detailed_rows = _check(_submission(report, "TEST-7"), "tortured_phrases")["result"]["supporting_data"]
            self.assertTrue(
                any(
                    row["validation_status"] == "rejected"
                    for row in detailed_rows
                )
            )

            workbook = load_workbook(result.output_paths["workbook"], read_only=True)
            ws = workbook["Check Detail"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            self.assertIn("Tortured Phrases - Evidence", headers)
            summary_headers = [cell.value for cell in next(workbook["All Abstracts"].iter_rows(max_row=1))]
            self.assertIn("Overall Risk", summary_headers)
            self.assertIn("tortured_rejected_count", result.abstract_summary_rows[0])

    def test_default_pipeline_leaves_validation_columns_blank(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir_str:
            temp_dir = Path(temp_dir_str)
            input_dir = temp_dir / "xmls"
            output_dir = temp_dir / "outputs"
            input_dir.mkdir()
            _write_temp_file(
                input_dir,
                "sample.xml",
                """
                <article article-type="Original Research">
                  <front>
                    <journal-meta><journal-title-group><journal-title>Test Journal</journal-title></journal-title-group></journal-meta>
                    <article-meta>
                      <article-id pub-id-type="manuscript">TEST-8</article-id>
                      <article-title>Default run</article-title>
                      <abstract><p>The model uses a nervous network in the abstract to ensure one finding exists.</p></abstract>
                      <history><date date-type="accepted"><year>2026</year></date></history>
                    </article-meta>
                  </front>
                </article>
                """,
            )
            dict_path = temp_dir / "dict.csv"
            dict_path.write_text(
                "Fingerprint - Tortured Phrase,Expected Text,Nb Retrieved Papers\n\"nervous network\",\"neural network\",\"1\"\n",
                encoding="utf-8",
            )

            result = run_default_pipeline(input_dir=input_dir, tortured_dictionary_path=dict_path, output_dir=output_dir)

            self.assertTrue(all(finding.validation_status == "" for finding in result.findings))
            self.assertTrue(all(finding.validation_reason == "" for finding in result.findings))
            self.assertTrue(all(finding.validated_by == "" for finding in result.findings))

    def _temp_xml(self, content: str) -> Path:
        temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xml")
        path = Path(temp.name)
        temp.close()
        path.write_text(textwrap.dedent(content).strip(), encoding="utf-8")
        return path


if __name__ == "__main__":
    unittest.main()
