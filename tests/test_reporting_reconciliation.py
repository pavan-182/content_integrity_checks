"""System-contract tests: JSON and Excel must tell exactly the same story.

These tests run the pipeline once and verify that the JSON (machine truth) and
the Excel workbook (editor view) are both projections of the same authoritative
result state, with no independently recalculated risk/status/flag values.
"""

from __future__ import annotations

import json
import tempfile
import textwrap
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from content_integrity.models import INACTIVE_VALIDATION_STATUSES
from content_integrity.pipeline import run_default_pipeline


ROOT = Path(__file__).resolve().parents[1]


def _write_temp_file(directory: Path, name: str, content: str) -> Path:
    path = directory / name
    path.write_text(textwrap.dedent(content).strip(), encoding="utf-8")
    return path


_TEMPLATE_ABSTRACT = (
    "Background: Patients with advanced disease entered this prospective study. "
    "Methods: The primary endpoint was independently assessed response rate. "
    "Results: Treatment produced durable responses and improved survival in the "
    "enrolled population. Conclusion: These findings support additional controlled "
    "investigation."
)


def _article_xml(record_id: str, title: str, abstract: str) -> str:
    return f"""
    <article article-type="Original Research">
      <front>
        <journal-meta><journal-title-group><journal-title>Test Journal</journal-title></journal-title-group></journal-meta>
        <article-meta>
          <article-id pub-id-type="manuscript">{record_id}</article-id>
          <article-title>{title}</article-title>
          <abstract><p>{abstract}</p></abstract>
          <history><date date-type="accepted"><year>2026</year></date></history>
        </article-meta>
      </front>
    </article>
    """


def _sheet_rows(worksheet) -> list[dict[str, object]]:
    # xlsx has no distinct empty-string cell type: openpyxl reads a blank cell
    # back as None even though "" was written, so normalize None -> "" here
    # rather than treating it as a JSON/Excel reconciliation mismatch.
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows)
    return [{key: ("" if value is None else value) for key, value in zip(headers, row)} for row in rows]


def _is_active(validation_status: str) -> bool:
    return (validation_status or "").strip().lower() not in INACTIVE_VALIDATION_STATUSES


class ReportingReconciliationTests(unittest.TestCase):
    """One pipeline run, checked against both of its own output artifacts."""

    @classmethod
    def setUpClass(cls) -> None:
        cls._tmp = tempfile.TemporaryDirectory()
        temp_dir = Path(cls._tmp.name)
        input_dir = temp_dir / "xmls"
        output_dir = temp_dir / "outputs"
        input_dir.mkdir()

        _write_temp_file(input_dir, "cluster_a.xml", _article_xml("TPL-A", "Template A", _TEMPLATE_ABSTRACT))
        _write_temp_file(input_dir, "cluster_b.xml", _article_xml("TPL-B", "Template B", _TEMPLATE_ABSTRACT))
        _write_temp_file(
            input_dir,
            "tortured.xml",
            _article_xml("TP-1", "Tortured phrase example", "The model uses a nervous network."),
        )
        _write_temp_file(input_dir, "clean.xml", _article_xml("CLEAN-1", "Clean abstract", "Nothing notable here."))

        dict_path = _write_temp_file(
            temp_dir,
            "dict.csv",
            'Fingerprint - Tortured Phrase,Expected Text,Nb Retrieved Papers\n"nervous network","neural network","1"\n',
        )

        with patch(
            "content_integrity.pipeline.detect_numerical_contradictions",
            side_effect=RuntimeError("detector unavailable"),
        ):
            cls.result = run_default_pipeline(input_dir=input_dir, tortured_dictionary_path=dict_path, output_dir=output_dir)

        cls.report = json.loads(cls.result.output_paths["content_integrity_json"].read_text(encoding="utf-8"))
        cls.workbook = load_workbook(cls.result.output_paths["workbook"], read_only=True, data_only=True)

    @classmethod
    def tearDownClass(cls) -> None:
        cls.workbook.close()
        cls._tmp.cleanup()

    def test_abstract_counts_reconcile(self) -> None:
        excel_rows = _sheet_rows(self.workbook["Abstracts"])
        self.assertEqual(len(self.report["abstracts"]), len(excel_rows))
        self.assertEqual(
            {abstract["abstract_id"] for abstract in self.report["abstracts"]},
            {row["record_id"] for row in excel_rows},
        )

    def test_finding_ids_and_validation_state_reconcile(self) -> None:
        json_finding_ids = {finding["finding_id"] for finding in self.report["findings"]}
        excel_rows = _sheet_rows(self.workbook["Findings"])
        excel_finding_ids = {row["finding_id"] for row in excel_rows}
        self.assertEqual(json_finding_ids, excel_finding_ids)

        excel_by_id = {row["finding_id"]: row for row in excel_rows}
        for finding in self.report["findings"]:
            excel_row = excel_by_id[finding["finding_id"]]
            self.assertEqual(finding["validation_status"], excel_row["validation_status"])
            self.assertEqual(finding["severity"], excel_row["severity"])
            self.assertEqual(finding["active"], _is_active(excel_row["validation_status"]))

    def test_active_finding_count_reconciles(self) -> None:
        json_active_count = sum(1 for finding in self.report["findings"] if finding["active"])
        excel_active_count = sum(
            1 for row in _sheet_rows(self.workbook["Findings"]) if _is_active(row["validation_status"])
        )
        self.assertEqual(json_active_count, excel_active_count)
        self.assertGreater(json_active_count, 0)

    def test_review_required_and_risk_reconcile_per_record(self) -> None:
        excel_by_record = {row["record_id"]: row for row in _sheet_rows(self.workbook["Abstracts"])}
        json_review_required_ids = {
            abstract["abstract_id"] for abstract in self.report["abstracts"] if abstract["review_required"]
        }
        excel_review_required_ids = {
            record_id for record_id, row in excel_by_record.items() if row["review_required"] == "Yes"
        }
        self.assertEqual(json_review_required_ids, excel_review_required_ids)
        self.assertTrue(json_review_required_ids, "fixture must exercise at least one review-required record")

        for abstract in self.report["abstracts"]:
            excel_row = excel_by_record[abstract["abstract_id"]]
            self.assertEqual(abstract["overall_content_risk"], excel_row["overall_content_risk"])

    def test_operational_issues_reconcile(self) -> None:
        json_issues = {
            (issue["component"], issue["record_id"], issue["error_type"])
            for issue in self.report["operational_issues"]
        }
        excel_issues = {
            (row["component"], row["record_id"], row["error_type"])
            for row in _sheet_rows(self.workbook["Operational_Issues"])
        }
        self.assertEqual(json_issues, excel_issues)
        self.assertIn(("numerical_contradiction", "", "RuntimeError"), json_issues)

    def test_operational_failure_is_not_displayed_as_clear(self) -> None:
        for abstract in self.report["abstracts"]:
            check = abstract["checks"]["numerical_contradiction"]
            self.assertFalse(check["flagged"])
            self.assertTrue(check["operational_failure"])

    def test_canonical_template_pair_reconciles_with_directional_excel_rows(self) -> None:
        self.assertEqual(len(self.report["template_pairs"]), 1)
        pair = self.report["template_pairs"][0]
        self.assertEqual({pair["record_a"], pair["record_b"]}, {"TPL-A", "TPL-B"})

        excel_rows = [row for row in _sheet_rows(self.workbook["Template_Pairs"]) if row["pair_id"] == pair["pair_id"]]
        self.assertEqual(len(excel_rows), 2, "one canonical JSON pair must produce exactly two directional Excel rows")
        self.assertEqual(
            {(row["left_record_id"], row["right_record_id"]) for row in excel_rows},
            {("TPL-A", "TPL-B"), ("TPL-B", "TPL-A")},
        )

    def test_reporting_does_not_recompute_severity_or_risk(self) -> None:
        # The Findings/Abstracts sheets are trimmed presentation views, but every
        # value they show must come from the same Finding/abstract_summary_rows
        # objects serialized into JSON -- never a second calculation.
        finding = next(item for item in self.report["findings"] if item["detector_type"] == "tortured_phrase")
        excel_row = next(
            row for row in _sheet_rows(self.workbook["Findings"]) if row["finding_id"] == finding["finding_id"]
        )
        self.assertEqual(finding["severity"], excel_row["severity"])
        self.assertEqual(finding["rule_id"], excel_row["rule_id"])


class RejectedFindingReconciliationTests(unittest.TestCase):
    def test_rejected_finding_is_inactive_in_both_json_and_excel(self) -> None:
        class RejectingClient:
            def complete(self, *, system: str, user: str, max_tokens: int = 150, temperature: float = 0.0) -> str:
                return json.dumps({"status": "rejected", "reason": "Standard terminology, not a substitution."})

        with tempfile.TemporaryDirectory() as temp_dir_str:
            temp_dir = Path(temp_dir_str)
            input_dir = temp_dir / "xmls"
            input_dir.mkdir()
            _write_temp_file(
                input_dir,
                "sample.xml",
                _article_xml("REJ-1", "Validator example", "The model uses a nervous network in the background."),
            )
            dict_path = _write_temp_file(
                temp_dir,
                "dict.csv",
                'Fingerprint - Tortured Phrase,Expected Text,Nb Retrieved Papers\n"nervous network","neural network","1"\n',
            )

            with patch("content_integrity.pipeline.build_gpt_oss_client", return_value=RejectingClient()):
                result = run_default_pipeline(
                    input_dir=input_dir,
                    tortured_dictionary_path=dict_path,
                    output_dir=temp_dir / "outputs",
                    validate_llm=True,
                )

            report = json.loads(result.output_paths["content_integrity_json"].read_text(encoding="utf-8"))
            workbook = load_workbook(result.output_paths["workbook"], read_only=True, data_only=True)
            try:
                finding = next(item for item in report["findings"] if item["detector_type"] == "tortured_phrase")
                self.assertEqual(finding["validation_status"], "rejected")
                self.assertFalse(finding["active"])

                excel_row = next(
                    row for row in _sheet_rows(workbook["Findings"]) if row["finding_id"] == finding["finding_id"]
                )
                self.assertEqual(excel_row["validation_status"], "rejected")
                self.assertFalse(_is_active(excel_row["validation_status"]))

                abstract = report["abstracts"][0]
                self.assertEqual(abstract["overall_content_risk"], "None")
                excel_abstract = _sheet_rows(workbook["Abstracts"])[0]
                self.assertEqual(excel_abstract["overall_content_risk"], "None")
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
