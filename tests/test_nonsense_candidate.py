from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from content_integrity.detectors import build_tortured_rule_index, load_tortured_rules
from content_integrity.detectors.nonsense_candidate import NonsenseCandidateDetector
from content_integrity.models import Finding, ParsedRecord
from content_integrity.pipeline import run_default_pipeline
from content_integrity.xml_parser import parse_xml


ROOT = Path(__file__).resolve().parents[1]


PLANTED = {
    "ERBB2 directly metabolized pembrolizumab": "directly metabolized pembrolizumab",
    "BRCA1 physically swallowed nivolumab": "physically swallowed nivolumab",
    "PIK3CA digested trastuzumab": "digested trastuzumab",
}


class StubClient:
    def __init__(self) -> None:
        self.sentences: list[str] = []
        self.calls = 0

    def complete(self, *, system: str, user: str, max_tokens: int, temperature: float) -> str:
        self.calls += 1
        results = []
        for item in json.loads(user)["sentences"]:
            sentence = item["sentence"]
            self.sentences.append(sentence)
            phrase = next((value for marker, value in PLANTED.items() if marker in sentence), "")
            results.append(
                {
                    "id": item["id"],
                    "understandable": not bool(phrase),
                    "suspected_phrase": phrase,
                    "explanation": (
                        "The entities are connected by wording that has no coherent biomedical meaning."
                        if phrase
                        else "The sentence is understandable."
                    ),
                    "confidence": "high" if phrase else "medium",
                }
            )
        return json.dumps({"results": results})


class NonsenseCandidateTests(unittest.TestCase):
    def setUp(self) -> None:
        self.client = StubClient()
        self.detector = NonsenseCandidateDetector(self.client)

    def test_clean_sentences_do_not_create_findings(self) -> None:
        records = [
            ParsedRecord("a.xml", record_id="clean-1", abstract_sections=[{"section": "Results", "text": "ERBB2 expression predicted response to trastuzumab in patients with breast cancer."}]),
            ParsedRecord("b.xml", record_id="clean-2", abstract_sections=[{"section": "Methods", "text": "Participants completed symptom questionnaires before each scheduled clinic visit."}]),
            ParsedRecord("c.xml", record_id="clean-3", abstract_sections=[{"section": "Background", "text": "Background:"}]),
        ]

        self.assertTrue(all(not self.detector.detect(record) for record in records))
        self.assertTrue(any("ERBB2" in sentence for sentence in self.client.sentences))

    def test_three_eval_corpus_candidates_are_annotated_low_severity(self) -> None:
        paths = sorted((ROOT / "tests" / "fixtures" / "eval_corpus" / "positives").glob("nonsense_candidate_*.xml"))
        findings = [self.detector.detect(parse_xml(path)) for path in paths]

        self.assertEqual(len(findings), 3)
        self.assertTrue(all(len(items) == 1 for items in findings))
        self.assertTrue(all(items[0].check_type == "nonsense_candidate" for items in findings))
        self.assertTrue(all(items[0].severity == "low" and items[0].confidence > 0 for items in findings))
        self.assertTrue(all(items[0].evidence_snippet for items in findings))

    def test_multiple_candidates_remain_marked_for_review(self) -> None:
        record = ParsedRecord(
            "sample.xml",
            record_id="two-candidates",
            abstract_sections=[{
                "section": "Results",
                "text": (
                    "ERBB2 directly metabolized pembrolizumab into a survival receptor during tumor growth. "
                    "BRCA1 physically swallowed nivolumab and converted the antibody into cellular memory."
                ),
            }],
        )
        findings = self.detector.detect(record)

        self.assertEqual(len(findings), 2)
        self.assertTrue(all(finding.validation_status == "candidate" for finding in findings))

    def test_model_failure_is_not_silently_treated_as_no_candidate(self) -> None:
        class BrokenClient:
            def __init__(self) -> None:
                self.calls = 0

            def complete(self, **kwargs):
                self.calls += 1
                return "not-json"

        client = BrokenClient()
        detector = NonsenseCandidateDetector(client)
        record = ParsedRecord(
            "sample.xml",
            record_id="broken-model",
            abstract_sections=[{
                "section": "Results",
                "text": "ERBB2 directly metabolized pembrolizumab into a survival receptor during tumor growth.",
            }],
        )

        findings, stats = detector.detect_records([record])

        self.assertEqual(findings, [])
        self.assertEqual(stats.failed_sentence_record_ids, ["broken-model"])
        self.assertEqual(client.calls, 2)
        self.assertEqual(stats.retry_count, 1)

    def test_deterministic_routes_select_candidates_without_a_drug_name(self) -> None:
        rules = load_tortured_rules(ROOT / "🤷_tortured.csv", "wiley_tortured_seed_v1")
        detector = NonsenseCandidateDetector(self.client, rule_index=build_tortured_rule_index(rules))
        record = ParsedRecord(
            "sample.xml",
            record_id="routes",
            abstract_sections=[{
                "section": "Results",
                "text": (
                    "Tumour volume decreased steadily steadily across the treated cohort over twelve weeks. "
                    "Overall survival was estimated with the Kaplan-Meier method in all enrolled patients."
                ),
            }],
        )
        _, stats = detector.detect_records([record])

        self.assertEqual(stats.route_counts, {"repeated_content_word": 1})
        self.assertEqual(stats.candidate_sentence_count, 1)

    def test_identical_sentences_are_reviewed_once(self) -> None:
        text = "ERBB2 directly metabolized pembrolizumab into a survival receptor during tumor growth."
        records = [
            ParsedRecord(f"{index}.xml", record_id=f"dup-{index}", abstract_sections=[{"section": "Results", "text": text}])
            for index in range(3)
        ]
        findings, stats = self.detector.detect_records(records)

        self.assertEqual(stats.candidate_sentence_count, 1)
        self.assertEqual(self.client.calls, 1)
        self.assertEqual(len(findings), 1)

    def test_model_phrase_absent_from_the_source_sentence_is_rejected(self) -> None:
        class HallucinatingClient:
            def complete(self, *, system: str, user: str, max_tokens: int, temperature: float) -> str:
                return json.dumps({
                    "results": [
                        {
                            "id": item["id"],
                            "understandable": False,
                            "suspected_phrase": "a phrase that never appeared anywhere",
                            "explanation": "Fabricated evidence.",
                            "confidence": "high",
                        }
                        for item in json.loads(user)["sentences"]
                    ]
                })

        detector = NonsenseCandidateDetector(HallucinatingClient())
        record = ParsedRecord(
            "sample.xml",
            record_id="hallucinated",
            abstract_sections=[{
                "section": "Results",
                "text": "ERBB2 directly metabolized pembrolizumab into a survival receptor during tumor growth.",
            }],
        )
        findings, stats = detector.detect_records([record])

        self.assertEqual(findings, [])
        self.assertEqual(stats.failed_sentence_record_ids, ["hallucinated"])

    def test_known_tortured_matches_are_routed_away_from_novel_candidates(self) -> None:
        record = ParsedRecord(
            "sample.xml",
            record_id="known",
            abstract_sections=[{
                "section": "Results",
                "text": "ERBB2 directly metabolized pembrolizumab into a survival receptor during tumor growth.",
            }],
        )
        known = Finding(
            finding_id="F1",
            record_id="known",
            source_file="sample.xml",
            detector_type="tortured_phrase",
            category="tortured_phrase",
            matched_text="directly metabolized pembrolizumab",
            evidence_snippet="",
            section_or_field="Results",
            severity="medium",
            confidence=0.87,
            rule_id="TP-TEST",
        )
        findings, stats = self.detector.detect_records([record], {"known": [known]})

        self.assertEqual(findings, [])
        self.assertEqual(stats.known_fingerprint_suppressed_count, 1)
        self.assertEqual(self.client.calls, 0)

    def test_pipeline_flag_is_opt_in(self) -> None:
        fixture = ROOT / "tests" / "fixtures" / "eval_corpus" / "positives" / "nonsense_candidate_01.xml"
        with tempfile.TemporaryDirectory() as directory:
            input_dir = Path(directory) / "input"
            input_dir.mkdir()
            (input_dir / fixture.name).write_bytes(fixture.read_bytes())
            default_result = run_default_pipeline(input_dir, ROOT / "🤷_tortured.csv", Path(directory) / "default")
            with patch("content_integrity.pipeline.build_gpt_oss_client", return_value=StubClient()):
                enabled_result = run_default_pipeline(
                    input_dir,
                    ROOT / "🤷_tortured.csv",
                    Path(directory) / "enabled",
                    detect_nonsense_candidates=True,
                )
            report = json.loads(enabled_result.output_paths["content_integrity_json"].read_text())
            workbook = load_workbook(enabled_result.output_paths["workbook"], read_only=True, data_only=True)
            workbook_values = [
                str(cell.value).lower()
                for sheet in workbook.worksheets
                for row in sheet.iter_rows()
                for cell in row
                if cell.value is not None
            ]
            workbook.close()

        self.assertFalse(any(item.detector_type == "nonsense_candidate" for item in default_result.findings))
        self.assertTrue(any(item.detector_type == "nonsense_candidate" for item in enabled_result.findings))
        summary = enabled_result.abstract_summary_rows[0]
        self.assertEqual(summary["nonsense_candidate_count"], 1)
        self.assertEqual(summary["total_finding_count"], 0)
        self.assertEqual(summary["highest_severity"], "None")
        self.assertEqual(summary["overall_content_risk"], "None")
        self.assertEqual(summary["review_required"], "No")
        self.assertNotIn("nonsense_candidate", json.dumps(report).lower())
        self.assertFalse(any(
            "nonsense candidate" in value or "nonsense_candidate" in value
            for value in workbook_values
        ))


if __name__ == "__main__":
    unittest.main()
