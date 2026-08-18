from __future__ import annotations

import json
import tempfile
import unittest
from dataclasses import replace
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from content_integrity.detectors import built_in_llm_rules, detect_llm_trace_candidates
from content_integrity.detectors.llm_trace import candidate_to_finding
from content_integrity.detectors.llm_trace_fusion import (
    fuse_llm_trace_candidates,
    llm_reviewer_priority,
)
from content_integrity.detectors.llm_trace_semantic import (
    SYSTEM_PROMPT,
    detect_semantic_traces,
    validate_model_response,
)
from content_integrity.models import ParsedRecord, TraceTextBlock
from content_integrity.pipeline import run_default_pipeline
from content_integrity.rules import load_llm_trace_rules
from content_integrity.validators.llm_trace_validator import (
    LLMTraceValidator,
    apply_llm_trace_validation,
)
from content_integrity.xml_parser import parse_xml


ROOT = Path(__file__).resolve().parents[1]


class StubValidatorClient:
    model_name = "test/model"

    def __init__(self, status: str = "confirmed") -> None:
        self.status = status
        self.payloads: list[dict[str, object]] = []

    def complete(self, *, system: str, user: str, max_tokens: int, temperature: float) -> str:
        self.payloads.append(json.loads(user))
        return json.dumps({"status": self.status, "reason": "The context supports this decision."})


def semantic_payload(record_id: str, traces: list[dict[str, object]]) -> str:
    return json.dumps({"results": [{"record_id": record_id, "traces": traces}]})


class LosslessTraceTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.record = parse_xml(ROOT / "tests" / "fixtures" / "llm_trace_lossless.xml")

    def test_lossless_blocks_preserve_structure_and_inline_text(self) -> None:
        text = "\n\n".join(block.source_text for block in self.record.trace_text_blocks)
        self.assertIn("User:\nRewrite the following abstract.\n\nAssistant:", text)
        self.assertIn("### Background", text)
        self.assertIn("---", text)
        self.assertIn("```text\ncopied interface\n```", text)
        self.assertIn("Curly “quotation” and single ‘quotation’", text)
        self.assertIn("As an AI language model", text)
        self.assertIn("Repeated phrase appears here.", text)
        self.assertNotIn("\n", self.record.abstract_text)

    def test_deterministic_detection_uses_exact_blocks_and_context(self) -> None:
        candidates = detect_llm_trace_candidates(self.record, built_in_llm_rules())
        role_candidates = [item for item in candidates if item.mapped_rule_id == "LLM-026"]
        self.assertEqual([item.matched_text.strip() for item in role_candidates], ["User:", "Assistant:"])
        self.assertTrue(all(item.section_label == "Background" for item in role_candidates))
        self.assertTrue(all(item.source_text[item.source_start:item.source_end] == item.matched_text for item in candidates))
        self.assertFalse(any("The user requested" in item.matched_text for item in candidates))
        quoted = next(
            item
            for item in candidates
            if item.mapped_rule_id == "LLM-001" and item.section_label == "Conclusions"
        )
        finding = candidate_to_finding(quoted)
        self.assertEqual(quoted.context.quotation_context, "quoted")
        self.assertEqual(finding.validation_status, "pending")
        self.assertIn("The paper quoted the output", finding.evidence_snippet)


class CatalogueTests(unittest.TestCase):
    def test_catalogue_loads_and_is_shared_with_semantic_prompt(self) -> None:
        rules = load_llm_trace_rules()
        self.assertEqual([rule.rule_id for rule in rules], [rule.rule_id for rule in built_in_llm_rules()])
        self.assertTrue(all(rule.rule_id in SYSTEM_PROMPT for rule in rules))

    def test_catalogue_rejects_duplicate_invalid_and_missing_rules(self) -> None:
        base = """
rules:
  - rule_id: LLM-X
    category: ai_self_identification
    signal_level: strong
    severity: high
    priority_eligible: true
    requires_validation: false
    regex_patterns: ['(']
    semantic_description: bad regex
    positive_examples: []
    negative_examples: []
    version: 1
"""
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "rules.yaml"
            path.write_text(base, encoding="utf-8")
            with self.assertRaises(ValueError):
                load_llm_trace_rules(path)
            path.write_text("rules:\n  - rule_id: LLM-X\n", encoding="utf-8")
            with self.assertRaises(ValueError):
                load_llm_trace_rules(path)


class SemanticResponseTests(unittest.TestCase):
    def setUp(self) -> None:
        text = "A refined version of your abstract appears below."
        self.record = ParsedRecord(
            source_file="one.xml",
            record_id="R1",
            title="Exact title",
            abstract_text=text,
            abstract_sections=[{"section": "Background", "text": text}],
            trace_text_blocks=[
                TraceTextBlock("title", "Title", "title", 0, "Exact title"),
                TraceTextBlock("abstract", "Background", "paragraph", 0, text),
            ],
        )

    def _trace(self, **updates):
        trace = {
            "match_type": "semantic_variant",
            "mapped_rule_id": "LLM-009",
            "category": "response_preamble",
            "matched_text": "a refined version",
            "section_or_field": "Background",
            "confidence": 0.93,
            "reason": "The sentence introduces an edited response.",
        }
        trace.update(updates)
        return trace

    def test_plain_fenced_and_embedded_json_restore_exact_source_case(self) -> None:
        plain = semantic_payload("R1", [self._trace()])
        for raw in (plain, f"```json\n{plain}\n```", f"analysis before {plain} after"):
            candidates = validate_model_response(raw, [self.record], "B1", "model")
            self.assertEqual(candidates[0].matched_text, "A refined version")

    def test_novel_candidate_gets_occurrence_id(self) -> None:
        trace = self._trace(
            match_type="novel_pattern_candidate",
            mapped_rule_id="",
            category="prompt_leakage",
        )
        candidate = validate_model_response(
            semantic_payload("R1", [trace]), [self.record], "B1", "model"
        )[0]
        self.assertTrue(candidate.mapped_rule_id.startswith("LLM-NOVEL-OCC-"))
        self.assertEqual(candidate.match_type, "novel_pattern_candidate")
        apply_llm_trace_validation(
            [candidate],
            LLMTraceValidator(StubValidatorClient("uncertain")),
        )
        self.assertEqual(
            llm_reviewer_priority([candidate_to_finding(candidate)], built_in_llm_rules()),
            "Low",
        )

    def test_ambiguous_wrong_section_duplicate_and_invalid_schema_are_rejected(self) -> None:
        valid = semantic_payload("R1", [self._trace()])
        with self.assertRaises(ValueError):
            validate_model_response(f"{valid}\n{valid}", [self.record], "B1", "model")
        with self.assertRaises(ValueError):
            validate_model_response(
                semantic_payload("R1", [self._trace(section_or_field="Results")]),
                [self.record],
                "B1",
                "model",
            )

    def test_failed_record_does_not_discard_safe_split_record(self) -> None:
        second = ParsedRecord(
            source_file="two.xml",
            record_id="R2",
            abstract_text="Clean second record.",
            abstract_sections=[{"section": "Abstract", "text": "Clean second record."}],
        )

        class SplitClient:
            model_name = "test/split"

            def complete(self, *, system: str, user: str, max_tokens: int, temperature: float) -> str:
                records = json.loads(user)["records"]
                if len(records) > 1 or records[0]["record_id"] == "R2":
                    return "{}"
                return json.dumps(
                    {"results": [{"record_id": records[0]["record_id"], "traces": []}]}
                )

        candidates, stats = detect_semantic_traces(SplitClient(), [self.record, second])
        self.assertEqual(candidates, [])
        self.assertEqual(stats.failed_record_ids, ["R2"])
        self.assertEqual(stats.batch_failure_count, 1)
        with self.assertRaises(ValueError):
            validate_model_response(
                semantic_payload("R1", [self._trace(), self._trace()]),
                [self.record],
                "B1",
                "model",
            )
        with self.assertRaises(ValueError):
            validate_model_response(
                semantic_payload("R1", [self._trace(confidence=True)]),
                [self.record],
                "B1",
                "model",
            )


class FusionValidationPriorityTests(unittest.TestCase):
    def setUp(self) -> None:
        self.rules = built_in_llm_rules()
        self.record = ParsedRecord(
            source_file="one.xml",
            record_id="R1",
            abstract_text="Here is the revised abstract.",
            abstract_sections=[{"section": "Abstract", "text": "Here is the revised abstract."}],
        )

    def test_fusion_prefers_deterministic_overlap(self) -> None:
        deterministic = detect_llm_trace_candidates(self.record, self.rules)[0]
        semantic = validate_model_response(
            semantic_payload(
                "R1",
                [{
                    "match_type": "semantic_variant",
                    "mapped_rule_id": "LLM-009",
                    "category": "response_preamble",
                    "matched_text": "Here is the revised abstract",
                    "section_or_field": "Abstract",
                    "confidence": 0.99,
                    "reason": "This introduces a revision.",
                }],
            ),
            [self.record],
            "B1",
            "model",
        )[0]
        fused = fuse_llm_trace_candidates([semantic, deterministic])
        self.assertEqual(len(fused), 1)
        self.assertEqual(fused[0].match_type, "known_pattern")

    def test_fusion_keeps_overlapping_semantic_candidates_in_different_categories(self) -> None:
        text = "As an AI language model, here is the revised abstract."
        record = ParsedRecord(
            source_file="overlap.xml",
            record_id="OVERLAP",
            abstract_text=text,
            abstract_sections=[{"section": "Abstract", "text": text}],
        )
        candidates = validate_model_response(
            semantic_payload(
                "OVERLAP",
                [
                    {
                        "match_type": "semantic_variant",
                        "mapped_rule_id": "LLM-001",
                        "category": "ai_self_identification",
                        "matched_text": "As an AI language model",
                        "section_or_field": "Abstract",
                        "confidence": 0.95,
                        "reason": "The speaker identifies itself as an AI.",
                    },
                    {
                        "match_type": "novel_pattern_candidate",
                        "mapped_rule_id": "",
                        "category": "response_preamble",
                        "matched_text": "language model, here is the revised abstract",
                        "section_or_field": "Abstract",
                        "confidence": 0.91,
                        "reason": "The phrase introduces a response to a requester.",
                    },
                ],
            ),
            [record],
            "B1",
            "model",
        )

        self.assertEqual(
            {candidate.category for candidate in fuse_llm_trace_candidates(candidates)},
            {"ai_self_identification", "response_preamble"},
        )

    def test_fusion_drops_cross_category_relabel_without_new_words(self) -> None:
        record = ParsedRecord(
            source_file="interface.xml",
            record_id="INTERFACE",
            abstract_text="Regenerate response:",
            abstract_sections=[{"section": "Abstract", "text": "Regenerate response:"}],
        )
        deterministic = detect_llm_trace_candidates(record, self.rules)[0]
        semantic = validate_model_response(
            semantic_payload(
                "INTERFACE",
                [{
                    "match_type": "semantic_variant",
                    "mapped_rule_id": "LLM-017",
                    "category": "prompt_leakage",
                    "matched_text": "Regenerate response:",
                    "section_or_field": "Abstract",
                    "confidence": 0.95,
                    "reason": "This appears to be a leaked instruction.",
                }],
            ),
            [record],
            "B1",
            "model",
        )[0]

        fused = fuse_llm_trace_candidates([deterministic, semantic])

        self.assertEqual([(candidate.match_type, candidate.mapped_rule_id) for candidate in fused], [("known_pattern", "LLM-022")])

    def test_fusion_keeps_cross_category_semantic_candidate_with_new_words(self) -> None:
        text = "As an AI language model, I do not have access to a specific dataset."
        record = ParsedRecord(
            source_file="capability.xml",
            record_id="CAPABILITY",
            abstract_text=text,
            abstract_sections=[{"section": "Abstract", "text": text}],
        )
        deterministic = detect_llm_trace_candidates(record, self.rules)[0]
        semantic = validate_model_response(
            semantic_payload(
                "CAPABILITY",
                [{
                    "match_type": "semantic_variant",
                    "mapped_rule_id": "LLM-006",
                    "category": "capability_disclaimer",
                    "matched_text": text,
                    "section_or_field": "Abstract",
                    "confidence": 0.95,
                    "reason": "This disclaims access to a dataset.",
                }],
            ),
            [record],
            "B1",
            "model",
        )[0]

        fused = fuse_llm_trace_candidates([deterministic, semantic])

        self.assertEqual(
            {candidate.category for candidate in fused},
            {"ai_self_identification", "capability_disclaimer"},
        )

    def test_direct_conversion_uses_shared_validation_requirement(self) -> None:
        record = ParsedRecord(
            source_file="contextual.xml",
            record_id="CTX-DIRECT",
            abstract_text="Certainly, here is the requested revision.",
            abstract_sections=[{"section": "Abstract", "text": "Certainly, here is the requested revision."}],
        )
        candidate = detect_llm_trace_candidates(record, self.rules)[0]
        candidate.rule = replace(candidate.rule, requires_validation=False)

        finding = candidate_to_finding(candidate)

        self.assertEqual(candidate.rule.signal_level, "contextual")
        self.assertEqual(finding.validation_status, "pending")
        self.assertEqual(finding.review_status, "needs_validation")

    def test_selective_validation_and_priority_states(self) -> None:
        strong = detect_llm_trace_candidates(self.record, self.rules)[0]
        apply_llm_trace_validation([strong], LLMTraceValidator(StubValidatorClient()))
        self.assertEqual(strong.validation_status, "not_required")
        self.assertEqual(strong.validated_by, "")
        finding = candidate_to_finding(strong)
        self.assertEqual(llm_reviewer_priority([finding], self.rules), "High")

        markdown_record = ParsedRecord(
            source_file="md.xml",
            record_id="MD",
            abstract_text="### Background",
            abstract_sections=[{"section": "Abstract", "text": "### Background"}],
        )
        supporting = detect_llm_trace_candidates(markdown_record, self.rules)[0]
        apply_llm_trace_validation([supporting], None)
        supporting_finding = candidate_to_finding(supporting)
        self.assertEqual(supporting_finding.review_status, "supporting_only")
        self.assertEqual(llm_reviewer_priority([supporting_finding, supporting_finding], self.rules), "None")

        contextual_record = ParsedRecord(
            source_file="contextual.xml",
            record_id="CTX",
            abstract_text="Certainly, here is the requested revision.",
            abstract_sections=[{"section": "Abstract", "text": "Certainly, here is the requested revision."}],
        )
        contextual = detect_llm_trace_candidates(contextual_record, self.rules)[0]
        confirmed_client = StubValidatorClient("confirmed")
        apply_llm_trace_validation([contextual], LLMTraceValidator(confirmed_client))
        self.assertEqual(
            llm_reviewer_priority([candidate_to_finding(contextual)], self.rules),
            "Medium",
        )

        client = StubValidatorClient("rejected")
        rejected_candidate = detect_llm_trace_candidates(contextual_record, self.rules)[0]
        apply_llm_trace_validation([rejected_candidate], LLMTraceValidator(client))
        rejected = candidate_to_finding(rejected_candidate)
        self.assertEqual(rejected.review_status, "excluded_by_validation")
        self.assertEqual(llm_reviewer_priority([rejected], self.rules), "None")
        self.assertIn("containing_paragraph", client.payloads[0])


class SemanticExecutionTests(unittest.TestCase):
    @staticmethod
    def _record(record_id: str, text: str = "Clean record text for the batch.") -> ParsedRecord:
        return ParsedRecord(
            source_file=f"{record_id}.xml",
            record_id=record_id,
            abstract_text=text,
            abstract_sections=[{"section": "Abstract", "text": text}],
        )

    def test_oversized_record_fails_without_calling_the_model(self) -> None:
        class NeverCalledClient:
            model_name = "test/never"

            def __init__(self) -> None:
                self.calls = 0

            def complete(self, *, system: str, user: str, max_tokens: int, temperature: float) -> str:
                self.calls += 1
                return "{}"

        client = NeverCalledClient()
        oversized = self._record("BIG", "word " * 20000)
        candidates, stats = detect_semantic_traces(client, [oversized])

        self.assertEqual(candidates, [])
        self.assertEqual(client.calls, 0)
        self.assertEqual(stats.failed_record_ids, ["BIG"])
        self.assertEqual(stats.batch_failure_count, 1)
        self.assertEqual(stats.request_count, 0)

    def test_retries_and_splits_stay_bounded_and_are_counted(self) -> None:
        class AlwaysFailingClient:
            model_name = "test/failing"

            def __init__(self) -> None:
                self.calls = 0

            def complete(self, *, system: str, user: str, max_tokens: int, temperature: float) -> str:
                self.calls += 1
                return "not json"

        client = AlwaysFailingClient()
        records = [self._record("R1"), self._record("R2")]
        candidates, stats = detect_semantic_traces(client, records, max_records_per_batch=2)

        self.assertEqual(candidates, [])
        # One batch of two: 2 attempts, then one split into two single records at 2 attempts each.
        self.assertEqual(client.calls, 6)
        self.assertEqual(stats.request_count, 6)
        self.assertEqual(stats.retry_count, 3)
        self.assertEqual(sorted(stats.failed_record_ids), ["R1", "R2"])
        self.assertEqual(stats.batch_failure_count, 2)
        self.assertEqual(stats.model_id, "test/failing")
        self.assertEqual(stats.prompt_version, "llm_trace_semantic_v3")

    def test_batches_run_concurrently_within_the_configured_bound(self) -> None:
        import threading
        import time

        class ProbeClient:
            model_name = "test/probe"

            def __init__(self) -> None:
                self.lock = threading.Lock()
                self.in_flight = 0
                self.peak = 0

            def complete(self, *, system: str, user: str, max_tokens: int, temperature: float) -> str:
                with self.lock:
                    self.in_flight += 1
                    self.peak = max(self.peak, self.in_flight)
                time.sleep(0.05)
                with self.lock:
                    self.in_flight -= 1
                submitted = json.loads(user)["records"]
                return json.dumps(
                    {"results": [{"record_id": item["record_id"], "traces": []} for item in submitted]}
                )

        records = [self._record(f"R{index}") for index in range(6)]
        bounded = ProbeClient()
        _, bounded_stats = detect_semantic_traces(
            bounded, records, max_records_per_batch=1, max_concurrent_batches=2,
        )
        self.assertEqual(bounded_stats.batch_count, 6)
        self.assertEqual(bounded_stats.max_concurrent_batches, 2)
        self.assertLessEqual(bounded.peak, 2)

        parallel = ProbeClient()
        detect_semantic_traces(parallel, records, max_records_per_batch=1, max_concurrent_batches=4)
        self.assertGreater(parallel.peak, 1)

    def test_candidate_order_is_stable_under_concurrency(self) -> None:
        class OrderedClient:
            model_name = "test/ordered"

            def complete(self, *, system: str, user: str, max_tokens: int, temperature: float) -> str:
                submitted = json.loads(user)["records"]
                return json.dumps(
                    {
                        "results": [
                            {
                                "record_id": item["record_id"],
                                "traces": [{
                                    "match_type": "novel_pattern_candidate",
                                    "mapped_rule_id": "",
                                    "category": "prompt_leakage",
                                    "matched_text": "Clean record text",
                                    "section_or_field": "Abstract",
                                    "confidence": 0.8,
                                    "reason": "Probe trace for ordering.",
                                }],
                            }
                            for item in submitted
                        ]
                    }
                )

        records = [self._record(f"R{index}") for index in range(8)]
        candidates, _ = detect_semantic_traces(
            OrderedClient(), records, max_records_per_batch=1, max_concurrent_batches=4,
        )
        self.assertEqual(
            [candidate.record_id for candidate in candidates],
            [record.record_id for record in records],
        )


class PipelineIntegrationTests(unittest.TestCase):
    def test_semantic_discovery_is_integrated_and_reported(self) -> None:
        class SemanticClient:
            model_name = "test/semantic"

            def complete(self, *, system: str, user: str, max_tokens: int, temperature: float) -> str:
                payload = json.loads(user)
                results = []
                for record in payload["records"]:
                    results.append(
                        {
                            "record_id": record["record_id"],
                            "traces": [{
                                "match_type": "novel_pattern_candidate",
                                "mapped_rule_id": "",
                                "category": "prompt_leakage",
                                "matched_text": "I omitted the limitations as requested",
                                "section_or_field": "Conclusions",
                                "confidence": 0.89,
                                "reason": "The sentence refers to a prior editing instruction.",
                            }],
                        }
                    )
                return json.dumps({"results": results})

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            input_dir = root / "input"
            input_dir.mkdir()
            (input_dir / "one.xml").write_text(
                """<article><front><journal-meta><journal-title-group><journal-title>J</journal-title></journal-title-group></journal-meta><article-meta>
<article-id pub-id-type="manuscript">SEM-1</article-id><title-group><article-title>Study</article-title></title-group>
<abstract><sec><title>Conclusions</title><p>I omitted the limitations as requested.</p></sec></abstract>
<history><date date-type="accepted"><year>2026</year></date></history></article-meta></front></article>""",
                encoding="utf-8",
            )
            dictionary = root / "tortured.csv"
            dictionary.write_text(
                "Fingerprint - Tortured Phrase,Expected Text,Nb Retrieved Papers\n",
                encoding="utf-8",
            )
            with patch("content_integrity.pipeline.build_gpt_oss_client", return_value=SemanticClient()):
                result = run_default_pipeline(
                    input_dir=input_dir,
                    tortured_dictionary_path=dictionary,
                    output_dir=root / "output",
                    detect_llm_semantic=True,
                )

            semantic = next(item for item in result.findings if item.check_type == "novel_pattern_candidate")
            self.assertTrue(semantic.rule_id.startswith("LLM-NOVEL-OCC-"))
            self.assertEqual(semantic.validation_status, "pending")
            self.assertEqual(result.abstract_summary_rows[0]["llm_review_priority"], "Low")
            metadata = dict(result.run_metadata_rows)
            self.assertTrue(metadata["llm_semantic_enabled"])
            self.assertEqual(metadata["llm_novel_candidate_count"], 1)
            report = json.loads(result.output_paths["content_integrity_json"].read_text())
            submission = next(iter(report.values()))
            llm_check = next(item for item in submission["checks"] if item["check_name"] == "llm_response_trace")
            row = next(item for item in llm_check["result"]["supporting_data"] if item["check_type"] == "novel_pattern_candidate")
            self.assertEqual(row["matched_phrase"], "I omitted the limitations as requested")
            self.assertIn("validation_status", row)
            self.assertIn("rule_id", row)
            workbook = load_workbook(result.output_paths["workbook"], read_only=True)
            sheet = workbook["Check Detail"]
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows)
            excel_row = dict(zip(headers, next(rows)))
            self.assertEqual(excel_row["LLM response trace - Flag"], "Review")
            self.assertIn("I omitted the limitations as requested", excel_row["LLM response trace - Evidence"])

    def test_semantic_batch_failure_becomes_an_operational_issue(self) -> None:
        class BrokenSemanticClient:
            model_name = "test/broken"

            def complete(self, *, system: str, user: str, max_tokens: int, temperature: float) -> str:
                return "not a json response"

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            input_dir = root / "input"
            input_dir.mkdir()
            (input_dir / "one.xml").write_text(
                """<article><front><journal-meta><journal-title-group><journal-title>J</journal-title></journal-title-group></journal-meta><article-meta>
<article-id pub-id-type="manuscript">SEM-2</article-id><title-group><article-title>Study</article-title></title-group>
<abstract><sec><title>Conclusions</title><p>An ordinary conclusion sentence.</p></sec></abstract>
<history><date date-type="accepted"><year>2026</year></date></history></article-meta></front></article>""",
                encoding="utf-8",
            )
            dictionary = root / "tortured.csv"
            dictionary.write_text(
                "Fingerprint - Tortured Phrase,Expected Text,Nb Retrieved Papers\n",
                encoding="utf-8",
            )
            with patch(
                "content_integrity.pipeline.build_gpt_oss_client", return_value=BrokenSemanticClient()
            ):
                result = run_default_pipeline(
                    input_dir=input_dir,
                    tortured_dictionary_path=dictionary,
                    output_dir=root / "output",
                    detect_llm_semantic=True,
                )

            issue = next(
                item for item in result.operational_issues
                if item.component == "llm_response_trace_semantic"
            )
            self.assertEqual(issue.record_id, "SEM-2")
            self.assertEqual(issue.error_type, "model_failure")
            metadata = dict(result.run_metadata_rows)
            self.assertEqual(metadata["llm_semantic_failed_record_count"], 1)
            self.assertEqual(metadata["llm_semantic_batch_failure_count"], 1)
            self.assertEqual(metadata["llm_semantic_request_count"], 2)
            self.assertEqual(metadata["llm_semantic_retry_count"], 1)
            self.assertTrue(
                any(
                    row["warning_code"] == "llm_semantic_batch_failed"
                    for row in result.parse_warning_rows
                )
            )


if __name__ == "__main__":
    unittest.main()
