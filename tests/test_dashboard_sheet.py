from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from content_integrity.pipeline import run_default_pipeline


ROOT = Path(__file__).resolve().parents[1]


class DashboardSheetTests(unittest.TestCase):
    def test_dashboard_cards_and_queues_reconcile_with_all_abstracts(self) -> None:
        with tempfile.TemporaryDirectory() as output_dir:
            result = run_default_pipeline(
                input_dir=ROOT / "tests" / "fixtures" / "eval_corpus",
                tortured_dictionary_path=ROOT / "🤷_tortured.csv",
                output_dir=output_dir,
            )
            workbook = load_workbook(result.output_paths["workbook"], data_only=True)

        master = workbook["All Abstracts"]
        headers = [cell.value for cell in master[1]]
        rows = [dict(zip(headers, values)) for values in master.iter_rows(min_row=2, values_only=True)]
        dashboard = workbook["Dashboard"]
        self.assertEqual(dashboard["A4"].value, len(rows))
        self.assertEqual(dashboard["C4"].value, sum(row["Overall Risk"] in {"None", "Low"} for row in rows))
        self.assertEqual(dashboard["E4"].value, sum(row["Overall Risk"] == "Medium" for row in rows))
        self.assertEqual(dashboard["G4"].value, sum(row["Overall Risk"] == "High" for row in rows))
        self.assertEqual(
            len(rows),
            sum(workbook[name].max_row - 1 for name in ("High Risk Queue", "Moderate Risk Queue", "Low Risk Queue")),
        )


if __name__ == "__main__":
    unittest.main()
