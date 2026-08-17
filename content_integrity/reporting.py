from __future__ import annotations

import csv
import json
from collections import defaultdict
from collections.abc import Iterable, Sequence
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .utils import ensure_parent_dir, normalize_whitespace, to_pipe_string


# Bump when the top-level JSON contract (build_content_integrity_frontend_json)
# gains/removes/renames a field in a way a machine consumer must react to.
SCHEMA_VERSION = "1.0"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_FONT = Font(name="Calibri", size=11)
PAIR_COLUMNS = [
    "pair_id",
    "record_id",
    "matched_record_id",
    "source_file",
    "matched_source_file",
    "title",
    "matched_title",
    "primary_match_type",
    "supporting_match_types",
    "confidence",
    "severity",
    "matched_sections",
    "matched_sentence_count",
    "shared_text_coverage",
    "original_text_similarity",
    "masked_skeleton_similarity",
    "ngram_similarity",
    "high_value_section_similarity",
    "weighted_section_similarity",
    "variable_substitutions",
    "relationship_context",
    "pair_classification",
    "evidence_excerpt",
    "review_status",
    "editor_label",
    "editor_notes",
]

TRIAGE_CHECKS = [
    ("tortured_phrase", "tortured_phrase_flag", "Tortured Phrases"),
    ("llm_response_trace", "llm_trace_flag", "LLM Response Trace"),
    ("numerical_contradiction", "numerical_contradiction_flag", "Numerical Contradiction"),
    ("design_contradiction", "design_contradiction_flag", "Design Contradiction"),
    ("unverifiable_clinical_trial", "unverifiable_trial_flag", "Unverifiable Trial"),
    ("template", "template_flag", "Templating (Cross-Author)"),
]

AUTHORSHIP_CHECKS = [
    ("submission_volume", "Submission Volume"),
    ("author_count_deviation", "Author Count Deviation"),
    ("affiliance_relevance", "Affiliation Relevance"),
    ("author_network", "Author Network"),
    ("retraction_history", "Retraction History"),
]


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, list):
        if not value:
            return ""
        if all(isinstance(item, dict) for item in value):
            if all({"section", "text"}.issubset(item.keys()) for item in value):
                parts = []
                for item in value:
                    section = normalize_whitespace(item.get("section", ""))
                    text = normalize_whitespace(item.get("text", ""))
                    if section and text:
                        parts.append(f"{section}: {text}")
                    elif text:
                        parts.append(text)
                if parts:
                    return " || ".join(parts)
            return json.dumps(value, ensure_ascii=False, sort_keys=True)
        return " | ".join(normalize_whitespace(str(item)) for item in value if normalize_whitespace(str(item)))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value if isinstance(value, str) else str(value)


def _join_lines(parts: Iterable[str]) -> str:
    return " | ".join(part.strip() for part in parts if part and part.strip())


def _write_table(
    ws,
    rows: Sequence[dict[str, Any]],
    columns: Sequence[str],
    start_row: int = 1,
    start_col: int = 1,
    auto_filter: bool = True,
) -> tuple[int, int]:
    row_idx = start_row
    col_idx = start_col
    for offset, column in enumerate(columns, start=col_idx):
        cell = ws.cell(row=row_idx, column=offset, value=column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in rows:
        row_idx += 1
        for offset, column in enumerate(columns, start=col_idx):
            value = _stringify(row.get(column, ""))
            cell = ws.cell(row=row_idx, column=offset, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = THIN_FONT
    if auto_filter and rows:
        ws.auto_filter.ref = f"{ws.cell(start_row, start_col).coordinate}:{ws.cell(row=row_idx, column=col_idx + len(columns) - 1).coordinate}"
    return row_idx, col_idx + len(columns) - 1


def write_jsonl(path: str | Path, rows: Iterable[dict[str, Any]]) -> Path:
    resolved = ensure_parent_dir(path)
    with resolved.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True))
            handle.write("\n")
    return resolved


def write_json(path: str | Path, data: Any) -> Path:
    resolved = ensure_parent_dir(path)
    with resolved.open("w", encoding="utf-8") as handle:
        json.dump(data, handle, ensure_ascii=False, indent=2)
        handle.write("\n")
    return resolved


def _split_values(value: str, separator: str = " | ") -> list[str]:
    if not value:
        return []
    return [item.strip() for item in str(value).split(separator) if item.strip()]


def _split_semicolon_values(value: str) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in str(value).split(";") if item.strip()]


def _build_tortured_evidence(findings: Iterable[dict[str, Any]], *, include_validation_status: bool = False) -> str:
    """Shared by the frontend JSON and the xlsx workbook so both surfaces show the
    same 'matched phrase → expected term' evidence for tortured-phrase findings."""

    def _line(finding: dict[str, Any]) -> str:
        base = (
            f"{finding.get('matched_text', '')} → {finding['expected_term']}"
            if finding.get("expected_term") else str(finding.get("matched_text", ""))
        )
        if include_validation_status and finding.get("validation_status"):
            return f"{base} [{finding['validation_status']}]"
        return base

    return _join_lines(_line(finding) for finding in findings)


def _build_llm_evidence(findings: list[dict[str, Any]]) -> str:
    if not findings:
        return ""
    finding = findings[0]
    if finding.get("evidence_snippet"):
        return str(finding["evidence_snippet"])
    return f"Residual LLM response phrase detected: '{finding['matched_phrase']}'."


def _build_templating_evidence(findings: list[dict[str, Any]]) -> str:
    if not findings:
        return ""
    return str(findings[0].get("matching_text_evidence") or "")


_EXACT_TEXT_SIGNALS = {
    "exact_original_body",
    "exact_results_section",
    "substantial_shared_original_block",
    "distinctive_shared_text",
    "multiple_distinctive_sentences",
    "rare_exact_phrase",
    "substantial_shared_text",
    "partial_or_reordered_reuse",
}


def _as_value_set(value: Any) -> set[str]:
    if isinstance(value, (list, tuple, set)):
        return set(value)
    return set(_split_values(str(value or "")))


def _pair_matched_id(record_id: str, pair: dict[str, Any]) -> str:
    return str(pair["right_record_id"] if str(pair["left_record_id"]) == record_id else pair["left_record_id"])


def _pair_section_similarity(pair: dict[str, Any]) -> float:
    # Raw enriched-pair rows carry `strongest_masked_section_similarity`; the frontend-normalized
    # finding shape renames it to `strongest_section_similarity`. Support both here.
    value = pair.get("strongest_section_similarity", pair.get("strongest_masked_section_similarity", 0.0))
    return float(value or 0.0)


def _template_match_signals(pair: dict[str, Any]) -> dict[str, Any]:
    """Which template-reuse basis this pair was actually matched on, so reasoning text can name it."""
    direct = set(_split_values(str(pair.get("direct_evidence", ""))))
    detectors = _as_value_set(pair.get("detector_evidence"))
    primary = set(_split_values(str(pair.get("primary_evidence", ""))))
    routes = _as_value_set(pair.get("retrieval_routes"))
    combined = direct | detectors | primary
    return {
        "exact_signals": sorted(combined & _EXACT_TEXT_SIGNALS),
        "exact_text": bool(combined & _EXACT_TEXT_SIGNALS),
        "entity_structure": "entity_normalized_template" in combined,
        "title_structure": "title_template" in routes or "masked_title_template_with_original_support" in (direct | primary),
        "section_structure": bool(pair.get("strongest_section")) and _pair_section_similarity(pair) > 0,
    }


def _build_templating_reason(pair: dict[str, Any], matched_abstract_id: str, same_author_group: bool) -> str:
    matched = str(matched_abstract_id or "")
    if not matched:
        return ""
    author_group = "same author group" if same_author_group else "different author group"
    signals = _template_match_signals(pair)
    # Report the % reading that actually matches what triggered the pair: word-for-word text
    # share for exact reuse, structural share for entity/title/section-normalized matches.
    if signals["exact_text"]:
        text_pct = float(pair.get("original_body_similarity") or 0.0)
        return f"{text_pct:.0%} shared text with {matched}, {author_group}."
    if signals["entity_structure"]:
        structure_pct = float(pair.get("masked_body_similarity") or 0.0)
        text_pct = float(pair.get("original_body_similarity") or 0.0)
        return (
            f"{structure_pct:.0%} shared structure with {matched} once biomedical entities are normalized "
            f"({text_pct:.0%} shared original text), {author_group}."
        )
    if signals["title_structure"]:
        title_pct = float(pair.get("masked_title_similarity") or 0.0)
        return f"{title_pct:.0%} shared title structure with {matched}, {author_group}."
    if signals["section_structure"]:
        section_pct = _pair_section_similarity(pair)
        section = str(pair.get("strongest_section") or "text").lower()
        return f"{section_pct:.0%} shared structure in the {section} section with {matched}, {author_group}."
    return f"Strong template reuse detected with {matched}, {author_group}."


def _template_evidence_pair(record_id: str, pair: dict[str, Any], record_lookup: dict[str, Any]) -> dict[str, Any]:
    matched_id = _pair_matched_id(record_id, pair)
    left_id, right_id = str(pair["left_record_id"]), str(pair["right_record_id"])
    left_text = str(pair.get("left_matched_text") or "")
    right_text = str(pair.get("right_matched_text") or "")
    if not left_text or not right_text:
        for part in str(pair.get("matching_text_evidence") or "").split(" || "):
            if part.startswith(f"{left_id}: "):
                left_text = part[len(left_id) + 2:]
            elif part.startswith(f"{right_id}: "):
                right_text = part[len(right_id) + 2:]
    if record_id == right_id:
        left_text, right_text = right_text, left_text
    same_authors = _same_author_group(record_lookup[record_id], record_lookup[matched_id])
    reason = _build_templating_reason(pair, matched_id, same_authors)
    return {
        "pair_id": pair["pair_id"],
        "submitted_abstract_id": record_id,
        "submitted_title": record_lookup[record_id].title,
        "matched_abstract_id": matched_id,
        "matched_title": record_lookup[matched_id].title,
        "submitted_evidence": left_text,
        "matched_evidence": right_text,
        "section": pair.get("strongest_section", ""),
        "similarity": pair.get("strongest_masked_section_similarity", 0.0),
        "same_author_group": same_authors,
        "reason": reason,
    }


def _is_high_confidence(value: Any) -> bool:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value) >= 0.80
    return normalize_whitespace(str(value)).lower() in {"high", "very_high"}


def _same_author_group(left: Any, right: Any) -> bool:
    left_authors = {normalize_whitespace(value).lower() for value in getattr(left, "authors", []) if normalize_whitespace(value)}
    right_authors = {normalize_whitespace(value).lower() for value in getattr(right, "authors", []) if normalize_whitespace(value)}
    return bool(left_authors and right_authors and left_authors == right_authors)


def build_content_integrity_frontend_json(
    records: list[Any],
    findings: list[Any],
    enriched_pair_rows: list[dict[str, Any]],
    enriched_abstract_rows: list[dict[str, Any]],
    abstract_summary_rows: list[dict[str, Any]],
    operational_issues: list[Any],
    generated_at: str,
    git_revision: str,
    run_metadata: dict[str, Any] | None = None,
    template_family_rows: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    record_lookup = {record.record_id: record for record in records}
    template_abstract_lookup = {row["record_id"]: row for row in enriched_abstract_rows}
    result_lookup = {str(row["record_id"]): row for row in abstract_summary_rows}
    if set(result_lookup) != {str(record.record_id) for record in records}:
        raise ValueError("Authoritative abstract results must cover every record exactly once.")
    operational_issue_rows = sorted(
        (issue.to_dict() if hasattr(issue, "to_dict") else dict(issue) for issue in operational_issues),
        key=lambda issue: (str(issue.get("component", "")), str(issue.get("record_id", "")), str(issue.get("error_type", ""))),
    )
    operational_by_record: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for issue in operational_issue_rows:
        operational_by_record[str(issue.get("record_id", ""))].append(issue)
    findings_by_record: dict[str, list[Finding]] = defaultdict(list)
    all_findings_by_record: dict[str, list[Finding]] = defaultdict(list)
    for finding in findings:
        all_findings_by_record[finding.record_id].append(finding)
        if finding.detector_type not in {
            "tortured_phrase",
            "llm_response_trace",
            "numerical_contradiction",
            "design_contradiction",
            "unverifiable_clinical_trial",
        }:
            continue
        findings_by_record[finding.record_id].append(finding)

    reviewable_pairs: list[dict[str, Any]] = [
        pair for pair in enriched_pair_rows
        if str(pair.get("review_priority", "None")) != "None"
        and str(pair.get("pair_class", "")) in {"possible_template_reuse", "possible_related_duplicate"}
    ]
    reviewed_pairs_by_record: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for pair in reviewable_pairs:
        reviewed_pairs_by_record[str(pair["left_record_id"])].append(pair)
        reviewed_pairs_by_record[str(pair["right_record_id"])].append(pair)

    def _pair_sort_key(pair: dict[str, Any], record_id: str) -> tuple[int, float, str]:
        priority_rank = {"High": 3, "Medium": 2, "Low": 1, "None": 0}
        return (
            -priority_rank.get(str(pair.get("review_priority", "None")), 0),
            -float(pair.get("editorial_score") or 0.0),
            str(pair.get("left_record_id") if str(pair.get("left_record_id")) != record_id else pair.get("right_record_id", "")),
        )

    abstracts: list[dict[str, Any]] = []
    for record in sorted(records, key=lambda record: str(record.record_id)):
        record_id = str(record.record_id)
        tortured_findings = [
            finding for finding in findings_by_record.get(record_id, []) if finding.detector_type == "tortured_phrase"
        ]
        llm_findings = [
            finding for finding in findings_by_record.get(record_id, []) if finding.detector_type == "llm_response_trace"
        ]
        numerical_findings = [
            finding for finding in findings_by_record.get(record_id, []) if finding.detector_type == "numerical_contradiction"
        ]
        design_findings = [
            finding for finding in findings_by_record.get(record_id, []) if finding.detector_type == "design_contradiction"
        ]
        trial_findings = [
            finding for finding in findings_by_record.get(record_id, []) if finding.detector_type == "unverifiable_clinical_trial"
        ]
        template_pairs = reviewed_pairs_by_record.get(record_id, [])
        record_operational_issues = [
            *operational_by_record.get("", []),
            *operational_by_record.get(record_id, []),
        ]

        def _component_failed(*components: str) -> bool:
            return any(issue.get("component") in components for issue in record_operational_issues)

        result = result_lookup[record_id]
        flagged_tortured = result["tortured_phrase_flag"] == "Yes"
        flagged_llm = result["llm_trace_flag"] == "Yes"
        flagged_numerical = result["numerical_contradiction_flag"] == "Yes"
        flagged_design = result["design_contradiction_flag"] == "Yes"
        flagged_trial = result["unverifiable_trial_flag"] == "Yes"
        flagged_templating = result["template_flag"] == "Yes"
        flagged_checks: list[str] = []
        if flagged_tortured:
            flagged_checks.append("tortured_phrases")
        if flagged_llm:
            flagged_checks.append("llm_response_trace")
        if flagged_numerical:
            flagged_checks.append("numerical_contradiction")
        if flagged_design:
            flagged_checks.append("design_contradiction")
        if flagged_trial:
            flagged_checks.append("unverifiable_trial")
        if flagged_templating:
            flagged_checks.append("Templating (Cross-Author)")

        overall_risk = str(result["overall_content_risk"])
        review_required = result["review_required"] == "Yes"
        abstract_template_meta = template_abstract_lookup.get(record_id, {})
        template_family_id = abstract_template_meta.get("family_id", "")
        template_family_size = abstract_template_meta.get("family_size", 0) or 0
        ordered_template_pairs = sorted(template_pairs, key=lambda pair: _pair_sort_key(pair, record_id))
        top_template_pair = ordered_template_pairs[0] if ordered_template_pairs else None
        top_template_matched_id = _pair_matched_id(record_id, top_template_pair) if top_template_pair else ""
        top_template_same_author_group = (
            _same_author_group(record_lookup[record_id], record_lookup[top_template_matched_id])
            if top_template_pair else False
        )

        abstract_findings: dict[str, Any] = {
            "abstract_id": record_id,
            "doi": record.doi,
            "title": record.title,
            "corresponding_author": record.primary_author or None,
            "overall_risk": overall_risk,
            "overall_content_risk": overall_risk,
            "review_required": review_required,
            "review_reason": result["review_reason"],
            "why_flagged": result["review_reason"] or "No review required.",
            "flagged_checks": flagged_checks,
            "high_confidence_flags": 0,
            "corroborating_flags": 0,
            "finding_ids": [finding.finding_id for finding in all_findings_by_record.get(record_id, [])],
            "template_pair_ids": [pair["pair_id"] for pair in ordered_template_pairs],
            "operational_issues": record_operational_issues,
            "checks": {
                "tortured_phrases": {
                    "flagged": flagged_tortured,
                    "match_count": sum(finding.active for finding in tortured_findings),
                    "review_candidate": any(finding.review_candidate for finding in tortured_findings),
                    "operational_failure": any(finding.validation_failed for finding in tortured_findings) or _component_failed("tortured_phrase"),
                    "evidence": _build_tortured_evidence([
                        {
                            "finding_id": finding.finding_id,
                            "matched_text": finding.matched_text,
                            "expected_term": finding.expected_term,
                            "evidence_snippet": finding.evidence_snippet,
                            "section": finding.section_or_field,
                            "severity": finding.severity,
                            "confidence": finding.confidence,
                            "rule_id": finding.rule_id,
                            "validation_status": finding.validation_status or "not_validated",
                            "active": finding.active,
                            "review_candidate": finding.review_candidate,
                        }
                        for finding in tortured_findings
                    ]),
                    "findings": [
                        {
                            "finding_id": finding.finding_id,
                            "matched_phrase": finding.matched_text,
                            "expected_term": finding.expected_term,
                            "evidence_snippet": finding.evidence_snippet,
                            "section": finding.section_or_field,
                            "severity": finding.severity,
                            "confidence": finding.confidence,
                            "validation_status": finding.validation_status or "not_validated",
                            "validation_reason": finding.validation_reason or "",
                            "validated_by": finding.validated_by or "",
                            "rule_id": finding.rule_id,
                            "active": finding.active,
                            "review_candidate": finding.review_candidate,
                        }
                        for finding in tortured_findings
                    ],
                },
                "llm_response_trace": {
                    "flagged": flagged_llm,
                    "match_count": sum(finding.active for finding in llm_findings),
                    "review_candidate": any(finding.review_candidate for finding in llm_findings),
                    "operational_failure": any(finding.validation_failed for finding in llm_findings) or _component_failed("llm_response_trace", "llm_response_trace_semantic"),
                    "evidence": _build_llm_evidence([
                        {
                            "finding_id": finding.finding_id,
                            "check_type": finding.check_type,
                            "category": finding.category,
                            "matched_phrase": finding.matched_text,
                            "evidence_snippet": finding.evidence_snippet,
                            "section": finding.section_or_field,
                            "severity": finding.severity,
                            "confidence": finding.confidence,
                            "rule_id": finding.rule_id,
                            "validation_status": finding.validation_status or "not_validated",
                            "active": finding.active,
                            "review_candidate": finding.review_candidate,
                        }
                        for finding in llm_findings
                    ]),
                    "findings": [
                        {
                            "finding_id": finding.finding_id,
                            "check_type": finding.check_type,
                            "category": finding.category,
                            "matched_phrase": finding.matched_text,
                            "evidence_snippet": finding.evidence_snippet,
                            "section": finding.section_or_field,
                            "severity": finding.severity,
                            "confidence": finding.confidence,
                            "validation_status": finding.validation_status or "not_validated",
                            "validation_reason": finding.validation_reason or "",
                            "validated_by": finding.validated_by or "",
                            "rule_id": finding.rule_id,
                            "active": finding.active,
                            "review_candidate": finding.review_candidate,
                        }
                        for finding in llm_findings
                    ],
                },
                "numerical_contradiction": {
                    "flagged": flagged_numerical,
                    "match_count": sum(finding.active for finding in numerical_findings),
                    "review_candidate": any(finding.review_candidate for finding in numerical_findings),
                    "operational_failure": any(finding.validation_failed for finding in numerical_findings) or _component_failed("numerical_contradiction"),
                    "evidence": _join_lines(finding.evidence_snippet for finding in numerical_findings),
                    "findings": [finding.to_dict() for finding in numerical_findings],
                },
                "design_contradiction": {
                    "flagged": flagged_design,
                    "match_count": sum(finding.active for finding in design_findings),
                    "review_candidate": any(finding.review_candidate for finding in design_findings),
                    "operational_failure": any(finding.validation_failed for finding in design_findings) or _component_failed("design_contradiction"),
                    "evidence": _join_lines(finding.evidence_snippet for finding in design_findings),
                    "findings": [finding.to_dict() for finding in design_findings],
                },
                "unverifiable_trial": {
                    "flagged": flagged_trial,
                    "match_count": sum(finding.active for finding in trial_findings),
                    "review_candidate": any(finding.review_candidate for finding in trial_findings),
                    "operational_failure": any(finding.validation_failed for finding in trial_findings) or _component_failed("unverifiable_clinical_trial", "clinical_trial_registry"),
                    "evidence": _join_lines(finding.evidence_snippet for finding in trial_findings),
                    "findings": [finding.to_dict() for finding in trial_findings],
                },
                "templating": {
                    "label": "Templating (Cross-Author)",
                    "flagged": flagged_templating,
                    "match_count": len(ordered_template_pairs),
                    "review_candidate": False,
                    "operational_failure": _component_failed("exact_text_reuse", "entity_normalized_template"),
                    "evidence": _build_templating_evidence([
                        {
                            "matched_abstract_id": pair["right_record_id"] if str(pair["left_record_id"]) == record_id else pair["left_record_id"],
                            "strongest_section": pair.get("strongest_section", ""),
                            "strongest_section_similarity": pair.get("strongest_masked_section_similarity", 0.0),
                            "matching_text_evidence": pair.get("matching_text_evidence", ""),
                            "same_author_group": _same_author_group(record_lookup[record_id], record_lookup[str(pair["right_record_id"] if str(pair["left_record_id"]) == record_id else pair["left_record_id"])]),
                        }
                        for pair in ordered_template_pairs
                    ]),
                    "reason": (
                        _build_templating_reason(top_template_pair, top_template_matched_id, top_template_same_author_group)
                        if top_template_pair else ""
                    ),
                    "evidence_pairs": [
                        _template_evidence_pair(record_id, pair, record_lookup)
                        for pair in ordered_template_pairs
                    ],
                    "template_family_id": template_family_id,
                    "family_size": template_family_size,
                    "highest_review_priority": ordered_template_pairs[0]["review_priority"] if ordered_template_pairs else "None",
                    "findings": [
                        {
                            "pair_id": pair["pair_id"],
                            "matched_abstract_id": pair["right_record_id"] if str(pair["left_record_id"]) == record_id else pair["left_record_id"],
                            "matched_title": pair["right_title"] if str(pair["left_record_id"]) == record_id else pair["left_title"],
                            "pair_class": pair["pair_class"],
                            "confidence": pair.get("confidence", ""),
                            "review_priority": pair["review_priority"],
                            "editorial_score": pair["editorial_score"],
                            "primary_evidence": pair["primary_evidence"],
                            "supporting_evidence": _split_values(str(pair.get("supporting_evidence", ""))),
                            "contextual_evidence": _split_values(str(pair.get("contextual_evidence", ""))),
                            "direct_evidence": pair.get("direct_evidence", ""),
                            "detector_evidence": _split_values(str(pair.get("detector_evidence", ""))),
                            "retrieval_routes": _split_values(str(pair.get("retrieval_routes", ""))),
                            "evidence": pair.get("matching_text_evidence", ""),
                            "reason": _build_templating_reason(
                                pair,
                                _pair_matched_id(record_id, pair),
                                _same_author_group(record_lookup[record_id], record_lookup[_pair_matched_id(record_id, pair)]),
                            ),
                            "same_author_group": _same_author_group(record_lookup[record_id], record_lookup[_pair_matched_id(record_id, pair)]),
                            "strongest_section": pair.get("strongest_section", ""),
                            "masked_title_similarity": pair.get("masked_title_similarity", 0.0),
                            "original_title_similarity": pair.get("original_title_similarity", 0.0),
                            "masked_body_similarity": pair.get("masked_body_similarity", 0.0),
                            "original_body_similarity": pair.get("original_body_similarity", 0.0),
                            "strongest_section_similarity": pair.get("strongest_masked_section_similarity", 0.0),
                            "likely_substitutions": _split_semicolon_values(str(pair.get("likely_substitutions", ""))),
                            "context_interpretation": pair.get("context_interpretation", ""),
                            "shared_trial_ids": _split_values(str(pair.get("shared_trial_ids", ""))),
                            "shared_databases": _split_values(str(pair.get("shared_databases", ""))),
                            "shared_populations": _split_values(str(pair.get("shared_populations", ""))),
                            "shared_endpoints": _split_values(str(pair.get("shared_endpoints", ""))),
                            "explicit_companion_wording": bool(pair.get("explicit_companion_wording", False)),
                        }
                        for pair in ordered_template_pairs
                    ],
                },
            },
        }
        abstract_findings["high_confidence_flags"] = sum((
            any(finding.active and _is_high_confidence(finding.confidence) for finding in tortured_findings),
            any(finding.active and _is_high_confidence(finding.confidence) for finding in llm_findings),
            any(finding.active and _is_high_confidence(finding.confidence) for finding in numerical_findings),
            any(finding.active and _is_high_confidence(finding.confidence) for finding in design_findings),
            any(finding.active and _is_high_confidence(finding.confidence) for finding in trial_findings),
            flagged_templating and any(_is_high_confidence(pair.get("confidence", "")) for pair in ordered_template_pairs),
        ))
        abstracts.append(abstract_findings)

    total_submissions = len(abstracts)
    high_risk = sum(1 for item in abstracts if item["overall_risk"] == "High")
    moderate_risk = sum(1 for item in abstracts if item["overall_risk"] == "Medium")
    low_risk = sum(1 for item in abstracts if item["overall_risk"] == "Low")
    no_risk = sum(1 for item in abstracts if item["overall_risk"] == "None")
    requires_editor_judgement = sum(bool(item["review_required"]) for item in abstracts)
    return {
        "schema_version": SCHEMA_VERSION,
        "run": {
            "generated_at": generated_at,
            "git_revision": git_revision,
            "report_version": "content-integrity-frontend-v1",
            **(run_metadata or {}),
        },
        "summary": {
            "total_submissions": total_submissions,
            "high_risk": high_risk,
            "moderate_risk": moderate_risk,
            "low_risk": low_risk,
            "no_risk": no_risk,
            "requires_editor_judgement": requires_editor_judgement,
            "cleared_without_manual_review": total_submissions - requires_editor_judgement,
        },
        "abstracts": abstracts,
        "findings": [
            finding.to_dict()
            for finding in sorted(findings, key=lambda item: (item.record_id, item.detector_type, item.finding_id))
        ],
        "template_pairs": [
            {
                "pair_id": pair["pair_id"],
                "record_a": pair["left_record_id"],
                "record_b": pair["right_record_id"],
                "title_a": pair["left_title"],
                "title_b": pair["right_title"],
                "pair_class": pair["pair_class"],
                "review_priority": pair["review_priority"],
                "confidence": pair.get("confidence", ""),
                "editorial_score": pair.get("editorial_score", ""),
                "primary_evidence": pair.get("primary_evidence", ""),
                "supporting_evidence": _split_values(str(pair.get("supporting_evidence", ""))),
                "strongest_section": pair.get("strongest_section", ""),
                "masked_body_similarity": pair.get("masked_body_similarity", 0.0),
                "original_body_similarity": pair.get("original_body_similarity", 0.0),
                "likely_substitutions": _split_semicolon_values(str(pair.get("likely_substitutions", ""))),
                "context_interpretation": pair.get("context_interpretation", ""),
                "matching_text_evidence": pair.get("matching_text_evidence", ""),
                "family_id": pair.get("family_id", ""),
            }
            for pair in sorted(reviewable_pairs, key=lambda pair: str(pair["pair_id"]))
        ],
        "template_families": template_family_rows or [],
        "operational_issues": operational_issue_rows,
    }


_LEVEL_RANK = {"NONE": 0, "LOW": 1, "MEDIUM": 2, "HIGH": 3}


def _integration_level(value: Any) -> str:
    normalized = normalize_whitespace(str(value)).upper()
    return "LOW" if normalized in {"", "NONE"} else normalized


def _normalized_doi(value: Any) -> str:
    doi = normalize_whitespace(str(value or "")).lower()
    for prefix in ("https://doi.org/", "http://doi.org/", "doi:"):
        if doi.startswith(prefix):
            return doi[len(prefix):].strip()
    return doi


def _load_authorship_checks(path: str | Path | None) -> dict[str, dict[str, dict[str, str]]]:
    resolved = Path(path) if path else None
    if not resolved or not resolved.exists():
        return {}
    with resolved.open(encoding="utf-8") as handle:
        report = json.load(handle)
    if not isinstance(report, dict):
        return {}
    lookup: dict[str, dict[str, dict[str, str]]] = {}
    for doi, submission in report.items():
        if not isinstance(submission, dict):
            continue
        checks = {
            str(check.get("check_name", "")): {
                "level": str(check.get("result", {}).get("level", "")),
                "comment": str(check.get("result", {}).get("comment", "")),
            }
            for check in submission.get("checks", [])
            if isinstance(check, dict) and str(check.get("check_name", "")) in {name for name, _ in AUTHORSHIP_CHECKS}
        }
        if not checks:
            continue
        key = _normalized_doi(doi)
        if key:
            lookup[key] = checks
        abstract_id = str(submission.get("abstract_id", "")).strip()
        if abstract_id:
            lookup[abstract_id.lower()] = checks
    return lookup


def _finding_level(findings: list[dict[str, Any]], *, failed: bool = False) -> str:
    active = [finding for finding in findings if finding.get("active", True)]
    if not active:
        return "UNKNOWN" if failed else "LOW"
    return max(
        (_integration_level(finding.get("severity")) for finding in active),
        key=lambda level: _LEVEL_RANK.get(level, 0),
    )


def _integration_check(
    check_name: str,
    check_description: str,
    source: dict[str, Any],
    clear_comment: str,
    *,
    evidence_role: str | None = None,
) -> dict[str, Any]:
    findings = list(source.get("findings") or [])
    check = {
        "check_name": check_name,
        "check_description": check_description,
        "result": {
            "level": _finding_level(findings, failed=bool(source.get("operational_failure"))),
            "supporting_data": findings,
            "comment": str(source.get("evidence") or clear_comment),
        },
    }
    if evidence_role:
        check["evidence_role"] = evidence_role
    return check


def _template_sub_checks(pair: dict[str, Any]) -> list[dict[str, Any]]:
    priority = _integration_level(pair.get("review_priority"))
    signals = _template_match_signals(pair)
    exact_signals = signals["exact_signals"]
    entity_triggered = signals["entity_structure"]
    title_triggered = signals["title_structure"]
    section_triggered = signals["section_structure"]

    def result(level: str, data: list[dict[str, Any]], comment: str) -> dict[str, Any]:
        return {"level": level, "supporting_data": data, "comment": comment}

    original_body_pct = float(pair.get("original_body_similarity") or 0.0)
    masked_body_pct = float(pair.get("masked_body_similarity") or 0.0)
    masked_title_pct = float(pair.get("masked_title_similarity") or 0.0)
    original_title_pct = float(pair.get("original_title_similarity") or 0.0)
    section_pct = _pair_section_similarity(pair)
    section_name = str(pair.get("strongest_section") or "").strip()

    return [
        {
            "check_name": "exact_text_reuse",
            "check_description": "Detects distinctive original text shared by two submissions.",
            "evidence_role": "PRIMARY",
            "result": result(
                priority if exact_signals else "LOW",
                [{
                    "signals": exact_signals,
                    "original_body_similarity": pair.get("original_body_similarity", 0.0),
                    "evidence": pair.get("evidence", ""),
                }] if exact_signals else [],
                f"{original_body_pct:.0%} of the original abstract text is shared word-for-word between the two submissions."
                if exact_signals else "No shared original text between the two abstracts.",
            ),
        },
        {
            "check_name": "entity_normalized_template",
            "check_description": "Detects a shared writing skeleton after biomedical entities are normalized.",
            "evidence_role": "PRIMARY",
            "result": result(
                priority if entity_triggered else "LOW",
                [{
                    "masked_body_similarity": pair.get("masked_body_similarity", 0.0),
                    "original_body_similarity": pair.get("original_body_similarity", 0.0),
                    "likely_substitutions": pair.get("likely_substitutions", []),
                }] if entity_triggered else [],
                f"{masked_body_pct:.0%} shared writing structure once biomedical entities are normalized out, "
                f"with {original_body_pct:.0%} shared original text."
                if entity_triggered else "No shared writing structure detected after entity normalization.",
            ),
        },
        {
            "check_name": "title_template",
            "check_description": "Detects structurally similar titles after entity normalization.",
            "evidence_role": "SUPPORTING",
            "result": result(
                priority if title_triggered else "LOW",
                [{
                    "masked_title_similarity": pair.get("masked_title_similarity", 0.0),
                    "original_title_similarity": pair.get("original_title_similarity", 0.0),
                }] if title_triggered else [],
                f"{masked_title_pct:.0%} shared title structure, with {original_title_pct:.0%} shared original title text."
                if title_triggered else "No shared title structure detected.",
            ),
        },
        {
            "check_name": "section_similarity",
            "check_description": "Compares corresponding sections of the two abstracts.",
            "evidence_role": "SUPPORTING",
            "result": result(
                priority if section_triggered else "LOW",
                [{
                    "strongest_section": pair.get("strongest_section", ""),
                    "strongest_section_similarity": pair.get("strongest_section_similarity", 0.0),
                    "supporting_evidence": pair.get("supporting_evidence", []),
                }] if section_triggered else [],
                f"{section_name or 'Strongest'} section shares {section_pct:.0%} structural similarity between the abstracts."
                if section_triggered else "No section reached a meaningful structural similarity.",
            ),
        },
    ]


def build_integrated_content_integrity_json(canonical_report: dict[str, Any]) -> dict[str, Any]:
    """Project the authoritative report into the DOI-keyed cross-pipeline contract."""
    output: dict[str, Any] = {}
    hidden_finding_ids = {
        finding.get("finding_id")
        for finding in canonical_report.get("findings", [])
        if finding.get("detector_type") == "nonsense_candidate"
    }
    for abstract in canonical_report["abstracts"]:
        key = _normalized_doi(abstract.get("doi")) or str(abstract["abstract_id"])
        if key in output:
            raise ValueError(f"Duplicate content-integrity integration key: {key}")

        checks = abstract["checks"]
        template = checks["templating"]
        evidence_pairs = {item["pair_id"]: item for item in template.get("evidence_pairs", [])}
        pair_data = []
        for pair in template.get("findings", []):
            evidence_pair = evidence_pairs.get(pair["pair_id"], {})
            pair_data.append({
                "evidence_scope": "PAIR",
                "pair_id": pair["pair_id"],
                "matched_abstract_id": pair["matched_abstract_id"],
                "matched_title": pair.get("matched_title", ""),
                "classification": pair.get("pair_class", ""),
                "review_priority": _integration_level(pair.get("review_priority")),
                "editorial_score": pair.get("editorial_score", 0.0),
                "reason": pair.get("reason", ""),
                "same_author_group": bool(pair.get("same_author_group")),
                "sub_checks": _template_sub_checks(pair),
                "context": {
                    "shared_trial_ids": pair.get("shared_trial_ids", []),
                    "shared_databases": pair.get("shared_databases", []),
                    "shared_populations": pair.get("shared_populations", []),
                    "shared_endpoints": pair.get("shared_endpoints", []),
                    "explicit_companion_wording": pair.get("explicit_companion_wording", False),
                    "interpretation": pair.get("context_interpretation", ""),
                },
                "submitted_evidence": evidence_pair.get("submitted_evidence", ""),
                "matched_evidence": evidence_pair.get("matched_evidence", ""),
            })

        record_supporting_checks = [
            _integration_check(
                "numerical_contradiction",
                "Detects inconsistent percentages, counts, ranges, or other numerical claims.",
                checks["numerical_contradiction"],
                "No active numerical contradiction was detected.",
                evidence_role="CORROBORATING",
            ),
            _integration_check(
                "design_contradiction",
                "Detects mutually inconsistent study-design statements.",
                checks["design_contradiction"],
                "No active study-design contradiction was detected.",
                evidence_role="CORROBORATING",
            ),
            _integration_check(
                "unverifiable_clinical_trial",
                "Checks whether stated clinical-trial identifiers are valid and verifiable.",
                checks["unverifiable_trial"],
                "No unverifiable clinical-trial reference was detected.",
                evidence_role="CORROBORATING",
            ),
        ]
        template_supporting_data = [*pair_data, {
            "evidence_scope": "RECORD",
            "supporting_checks": record_supporting_checks,
        }]
        if template.get("template_family_id"):
            family = next(
                (
                    item for item in canonical_report.get("template_families", [])
                    if item.get("family_id") == template["template_family_id"]
                ),
                {},
            )
            template_supporting_data.append({
                "evidence_scope": "FAMILY",
                "family_id": template["template_family_id"],
                "family_size": template.get("family_size", 0),
                "representative_abstract_id": family.get("representative_record_id", ""),
                "member_abstract_ids": _split_values(str(family.get("member_ids", ""))),
            })

        template_level = (
            "UNKNOWN"
            if template.get("operational_failure") and not pair_data
            else _integration_level(template.get("highest_review_priority"))
        )
        triggered_checks = []
        if checks["tortured_phrases"].get("flagged") or checks["tortured_phrases"].get("review_candidate"):
            triggered_checks.append("Tortured Phrase")
        if checks["llm_response_trace"].get("flagged") or checks["llm_response_trace"].get("review_candidate"):
            triggered_checks.append("LLM Response Trace")
        if pair_data:
            triggered_checks.append("Template Detection")
        supporting_labels = {
            "numerical_contradiction": "Numerical Contradiction",
            "design_contradiction": "Design Contradiction",
            "unverifiable_clinical_trial": "Unverifiable Clinical Trial",
        }
        triggered_checks.extend(
            supporting_labels[check["check_name"]]
            for check in record_supporting_checks
            if check["result"]["supporting_data"]
        )
        summary_comment = ", ".join(triggered_checks) or "No active content-integrity detection."
        summary_data = {
            "overall_content_risk": abstract["overall_content_risk"].upper(),
            "review_required": abstract["review_required"],
            "flagged_checks": triggered_checks,
            "finding_ids": [
                finding_id for finding_id in abstract["finding_ids"]
                if finding_id not in hidden_finding_ids
            ],
            "template_pair_ids": abstract["template_pair_ids"],
            "operational_issues": abstract["operational_issues"],
        }
        output[key] = {
            "title": abstract["title"],
            "abstract_id": abstract["abstract_id"],
            "checks": [
                {
                    "check_name": "content_integrity_summary",
                    "check_description": "Authoritative aggregate result for all content-integrity checks.",
                    "result": {
                        "level": _integration_level(abstract["overall_content_risk"]),
                        "supporting_data": [summary_data],
                        "comment": summary_comment,
                    },
                },
                _integration_check(
                    "tortured_phrases",
                    "Detects known tortured or nonsensical scientific phrases.",
                    checks["tortured_phrases"],
                    "No active tortured phrase was detected.",
                ),
                _integration_check(
                    "llm_response_trace",
                    "Detects residual language associated with an LLM or conversational assistant response.",
                    checks["llm_response_trace"],
                    "No active LLM response trace was detected.",
                ),
                {
                    "check_name": "template_detection",
                    "check_description": "Detects suspicious text or writing-template reuse across submissions.",
                    "result": {
                        "level": template_level,
                        "supporting_data": template_supporting_data,
                        "comment": template.get("reason") or "No accepted template pair was detected.",
                    },
                },
            ],
        }
    return output


def write_csv(path: str | Path, rows: Sequence[dict[str, Any]], columns: Sequence[str]) -> Path:
    resolved = ensure_parent_dir(path)
    with resolved.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(columns))
        writer.writeheader()
        for row in rows:
            writer.writerow({column: _stringify(row.get(column, "")) for column in columns})
    return resolved


def write_workbook(
    path: str | Path,
    *,
    abstract_summary_rows: list[dict[str, Any]],
    findings_rows: list[dict[str, Any]],
    pair_rows: list[dict[str, Any]],
    operational_issue_rows: list[dict[str, Any]],
    run_metadata_rows: list[tuple[str, Any]],
    authorship_json_path: str | Path | None = None,
) -> Path:
    """Write the compact editor-triage workbook; JSON retains the diagnostic detail."""
    resolved = ensure_parent_dir(path)
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    authorship_checks_by_key = _load_authorship_checks(authorship_json_path)

    findings_by_record: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for finding in findings_rows:
        findings_by_record[str(finding.get("record_id", ""))].append(finding)
    pairs_by_record: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for pair in pair_rows:
        pairs_by_record[str(pair.get("left_record_id", ""))].append(pair)
    issues_by_record: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for issue in operational_issue_rows:
        issues_by_record[str(issue.get("record_id", ""))].append(issue)

    def evidence(record_id: str, detector: str) -> str:
        if detector == "template":
            return _join_lines(
                f"{pair.get('pair_id', '')}: {pair.get('right_record_id', '')} — "
                f"{pair.get('primary_evidence') or pair.get('matching_text_evidence') or pair.get('priority_reason', '')}"
                for pair in pairs_by_record.get(record_id, [])
            )
        if detector == "tortured_phrase":
            return _build_tortured_evidence(
                (item for item in findings_by_record.get(record_id, []) if item.get("detector_type") == detector),
                include_validation_status=True,
            )
        return _join_lines(
            f"{item.get('evidence_snippet', '')} [{item.get('validation_status')}]"
            if item.get("validation_status") else str(item.get("evidence_snippet", ""))
            for item in findings_by_record.get(record_id, [])
            if item.get("detector_type") == detector
        )

    def check_value(row: dict[str, Any], detector: str, flag: str) -> str:
        if row.get(flag) == "Yes":
            return "Y"
        if any(item.get("review_candidate") for item in findings_by_record.get(str(row["record_id"]), []) if item.get("detector_type") == detector):
            return "Review"
        return "N"

    def authorship_value(row: dict[str, Any], check_name: str) -> str:
        keys = (
            _normalized_doi(row.get("doi")),
            str(row.get("record_id", "")).lower(),
        )
        for key in keys:
            if not key:
                continue
            check = authorship_checks_by_key.get(key, {}).get(check_name)
            if check:
                return check.get("level", "")
        return ""

    def authorship_evidence(row: dict[str, Any], check_name: str) -> str:
        keys = (
            _normalized_doi(row.get("doi")),
            str(row.get("record_id", "")).lower(),
        )
        for key in keys:
            if not key:
                continue
            check = authorship_checks_by_key.get(key, {}).get(check_name)
            if check:
                return check.get("comment", "")
        return ""

    def why_flagged_text(row: dict[str, Any], values: dict[str, str]) -> str:
        reasons: list[str] = []
        for _, flag, label in TRIAGE_CHECKS:
            if values.get(label) != "N":
                reasons.append(label)
        for check_name, label in AUTHORSHIP_CHECKS:
            level = authorship_value(row, check_name)
            if level in {"MEDIUM", "HIGH"}:
                reasons.append(label)
        if reasons:
            return ", ".join(reasons)
        if any(issues_by_record.get(str(row["record_id"]), [])) or issues_by_record.get("", []):
            return "Operational Issues"
        return "No active integrity findings."

    master_rows: list[dict[str, Any]] = []
    check_rows: list[dict[str, Any]] = []
    for row in abstract_summary_rows:
        record_id = str(row["record_id"])
        values = {label: check_value(row, detector, flag) for detector, flag, label in TRIAGE_CHECKS}
        record_issues = [*issues_by_record.get("", []), *issues_by_record.get(record_id, [])]
        why = why_flagged_text(row, values)
        master = {
            "Abstract ID": record_id,
            "Title (short)": str(row.get("title", ""))[:120],
            "Corresponding Author": row.get("primary_author") or str(row.get("authors", "")).split(" | ")[0],
            "Overall Risk": row.get("overall_content_risk", "None"),
            "Why Flagged (plain English)": why,
            **values,
            **{label: authorship_value(row, check_name) for check_name, label in AUTHORSHIP_CHECKS},
            "Operational Issues": "Y" if record_issues else "N",
            "Finding Count": row.get("active_finding_count", 0),
            "Review Required": row.get("review_required", "No"),
        }
        master_rows.append(master)
        detail = {
            "Abstract ID": master["Abstract ID"],
            "Title (short)": master["Title (short)"],
            "Corresponding Author": master["Corresponding Author"],
        }
        for detector, _, label in TRIAGE_CHECKS:
            detail[f"{label} - Flag"] = values[label]
            detail[f"{label} - Evidence"] = evidence(record_id, detector)
        for check_name, label in AUTHORSHIP_CHECKS:
            detail[f"{label} - Level"] = authorship_value(row, check_name)
            detail[f"{label} - Evidence"] = authorship_evidence(row, check_name)
        detail["Operational Issues - Flag"] = master["Operational Issues"]
        detail["Operational Issues - Evidence"] = _join_lines(
            f"{item.get('component', '')}: {item.get('message', '')}" for item in record_issues
        )
        check_rows.append(detail)

    risk_groups = {
        "High": [row for row in master_rows if row["Overall Risk"] == "High"],
        "Medium": [row for row in master_rows if row["Overall Risk"] == "Medium"],
        "Low": [row for row in master_rows if row["Overall Risk"] not in {"High", "Medium"}],
    }
    for label, rows in risk_groups.items():
        for rank, row in enumerate(rows, 1):
            row[f"{label} Risk Rank"] = rank
    for row in master_rows:
        for label in risk_groups:
            row.setdefault(f"{label} Risk Rank", "")

    def style_table(ws, rows: list[dict[str, Any]], columns: list[str], *, detail: bool = False) -> None:
        _write_table(ws, rows, columns)
        ws.freeze_panes = "D2" if len(columns) > 7 else "B2"
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 42
        for column, name in enumerate(columns, 1):
            color = "8EA9DB" if column <= 3 else ("C0504D" if detail and name.endswith("- Flag") else "1F4E78")
            ws.cell(1, column).fill = PatternFill("solid", fgColor=color)
            ws.cell(1, column).font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            ws.cell(1, column).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_index in range(2, ws.max_row + 1):
            ws.row_dimensions[row_index].height = 34
            for cell in ws[row_index]:
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(horizontal="center" if cell.column <= 4 else "left", vertical="center", wrap_text=True)
        widths = {1: 15, 2: 38, 3: 22, 4: 14, 5: 48}
        for column in range(1, len(columns) + 1):
            ws.column_dimensions[ws.cell(1, column).column_letter].width = widths.get(column, 18 if "Flag" in columns[column - 1] else 42)

    # Dashboard mirrors the supplied editor workbook: four cards plus compact triage summaries.
    dashboard.sheet_view.showGridLines = False
    dashboard.merge_cells("A1:H1")
    dashboard["A1"] = "Editor Triage — Submission Overview"
    dashboard["A1"].font = Font(name="Arial", size=14, bold=True, color="1F4E78")
    cards = [
        ("A3:B3", "A4:B5", "Total Submissions", len(master_rows), "D9E1F2"),
        ("C3:D3", "C4:D5", "No / Low Risk", len(risk_groups["Low"]), "E2F0D9"),
        ("E3:F3", "E4:F5", "Moderate Risk", len(risk_groups["Medium"]), "FCE4D6"),
        ("G3:H3", "G4:H5", "High Risk", len(risk_groups["High"]), "F4CCCC"),
    ]
    for label_range, value_range, label, value, color in cards:
        dashboard.merge_cells(label_range)
        dashboard.merge_cells(value_range)
        label_cell = dashboard[label_range.split(":")[0]]
        value_cell = dashboard[value_range.split(":")[0]]
        label_cell.value, value_cell.value = label, value
        label_cell.font = Font(name="Arial", size=10, color="595959")
        value_cell.font = Font(name="Arial", size=22, bold=True)
        label_cell.alignment = value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        value_cell.fill = PatternFill("solid", fgColor=color)
    dashboard.merge_cells("A7:H7")
    dashboard["A7"] = f"{sum(row['Review Required'] == 'Yes' for row in master_rows)} submissions require editor judgement"
    dashboard["A7"].font = Font(name="Arial", size=11, bold=True, color="1F4E78")
    dashboard["A7"].alignment = Alignment(horizontal="center")
    dashboard.merge_cells("A10:H10")
    dashboard["A10"] = "Active checks"
    dashboard["A10"].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    dashboard["A10"].fill = HEADER_FILL
    check_summary = [{"Check": label, "Flagged Abstracts": sum(row[label] == "Y" for row in master_rows)} for _, _, label in TRIAGE_CHECKS]
    _write_table(dashboard, check_summary, ["Check", "Flagged Abstracts"], start_row=11, auto_filter=False)
    dashboard.column_dimensions["A"].width = 32
    for column in "BCDEFGH":
        dashboard.column_dimensions[column].width = 14

    queue_columns = ["Rank", "Abstract ID", "Title (short)", "Corresponding Author", "Why Flagged (plain English)"]
    for sheet_name, key in (("High Risk Queue", "High"), ("Moderate Risk Queue", "Medium"), ("Low Risk Queue", "Low")):
        rows = [{"Rank": index, **row} for index, row in enumerate(risk_groups[key], 1)]
        ws = workbook.create_sheet(sheet_name)
        style_table(ws, rows, queue_columns)

    master_columns = [
        "Abstract ID", "Title (short)", "Corresponding Author", "Overall Risk", "Why Flagged (plain English)",
        *(label for _, _, label in TRIAGE_CHECKS),
        *(label for _, label in AUTHORSHIP_CHECKS),
        "Operational Issues", "Finding Count", "Review Required", "High Risk Rank", "Medium Risk Rank", "Low Risk Rank",
    ]
    ws = workbook.create_sheet("All Abstracts")
    style_table(ws, master_rows, master_columns)
    for row_index in range(2, ws.max_row + 1):
        risk_cell = ws.cell(row_index, 4)
        risk_cell.fill = PatternFill("solid", fgColor={"High": "F4CCCC", "Medium": "FCE4D6", "Low": "E2F0D9", "None": "E2F0D9"}.get(str(risk_cell.value), "FFFFFF"))

    detail_columns = ["Abstract ID", "Title (short)", "Corresponding Author"]
    for _, _, label in TRIAGE_CHECKS:
        detail_columns.extend((f"{label} - Flag", f"{label} - Evidence"))
    for _, label in AUTHORSHIP_CHECKS:
        detail_columns.extend((f"{label} - Level", f"{label} - Evidence"))
    detail_columns.extend(("Operational Issues - Flag", "Operational Issues - Evidence"))
    ws = workbook.create_sheet("Check Detail")
    style_table(ws, check_rows, detail_columns, detail=True)

    ws = workbook.create_sheet("How This Works")
    ws.sheet_view.showGridLines = False
    instructions = [
        "Editor Triage Workbook — How This Works",
        "",
        "Start with Dashboard, then use the High and Moderate Risk queues. All Abstracts is the complete authoritative list.",
        "1. Risk and review values come from the pipeline decision engine; this workbook does not recalculate them.",
        "2. Check Detail contains the submitted evidence and validation state behind each check.",
        "3. 'Review' means an inactive candidate requires judgement; it is not confirmed risk evidence.",
        "4. Authorship checks come from final_json.json; their displayed value is result.level and their evidence is the check comment.",
        "5. Operational issues mean a check did not complete; they are not scientific findings.",
        "6. The companion content_integrity_results.json remains the complete machine-readable report.",
        "",
        "Run metadata",
    ]
    for row_index, value in enumerate(instructions, 1):
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=6)
        cell = ws.cell(row_index, 1, value=value)
        cell.font = Font(name="Arial", size=14 if row_index == 1 else 10, bold=row_index in {1, 10}, color="1F4E78")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_index].height = 28
    metadata_start = len(instructions) + 1
    for offset, (key, value) in enumerate(run_metadata_rows):
        row_index = metadata_start + offset
        ws.cell(row_index, 1, key).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row_index, 2, _stringify(value)).font = Font(name="Arial", size=10)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 90

    workbook.save(resolved)
    return resolved
