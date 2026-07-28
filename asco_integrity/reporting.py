from __future__ import annotations

import csv
import json
from collections import Counter
from collections.abc import Iterable, Sequence
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .utils import ensure_parent_dir, normalize_whitespace, to_pipe_string


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_FONT = Font(name="Calibri", size=11)

FINDINGS_COLUMNS = [
    "finding_id",
    "record_id",
    "source_file",
    "detector_type",
    "check_type",
    "category",
    "matched_text",
    "expected_term",
    "evidence_snippet",
    "section_or_field",
    "severity",
    "confidence",
    "validation_status",
    "validation_reason",
    "validated_by",
    "rule_id",
    "template_cluster_id",
    "cluster_size",
    "similar_record_ids",
    "similarity_score",
    "cluster_severity",
    "shared_skeleton_excerpt",
    "metadata_context",
    "template_pattern_type",
    "original_text_similarity",
    "masked_skeleton_similarity",
    "ngram_similarity",
    "weighted_section_similarity",
    "section_similarities",
    "variable_substitutions",
    "exclusion_reason",
    "pair_id",
    "matched_record_id",
    "matched_source_file",
    "title",
    "matched_title",
    "primary_match_type",
    "supporting_match_types",
    "matched_sections",
    "matched_sentence_count",
    "shared_text_coverage",
    "high_value_section_similarity",
    "relationship_context",
    "review_status",
]

REVIEW_FINDINGS_COLUMNS = [
    "finding_id",
    "detector_type",
    "check_type",
    "record_id",
    "matched_record_id",
    "title",
    "matched_title",
    "evidence_snippet",
    "matched_sentence_count",
    "shared_text_coverage",
    "relationship_context",
    "severity",
    "confidence",
    "review_status",
]

PAIR_COLUMNS = [
    "pair_id",
    "record_id",
    "matched_record_id",
    "source_file",
    "matched_source_file",
    "title",
    "matched_title",
    "primary_match_type",
    "supporting_match_types",
    "confidence",
    "severity",
    "matched_sections",
    "matched_sentence_count",
    "shared_text_coverage",
    "original_text_similarity",
    "masked_skeleton_similarity",
    "ngram_similarity",
    "high_value_section_similarity",
    "weighted_section_similarity",
    "variable_substitutions",
    "relationship_context",
    "evidence_excerpt",
    "review_status",
]

FAMILY_COLUMNS = [
    "template_family_id",
    "member_count",
    "family_confidence",
    "representative_record_id",
    "template_pattern_type",
    "matched_sections",
    "edge_density",
    "median_pair_confidence",
    "changed_entity_types",
    "member_ids",
    "shared_skeleton_excerpt",
    "medoid_verification_passed",
]

DICTIONARY_COLUMNS = [
    "detector_type",
    "rule_id",
    "category",
    "pattern",
    "matched_phrase",
    "expected_term",
    "severity",
    "confidence",
    "retrieved_papers",
    "source",
]

WARNING_COLUMNS = [
    "source_file",
    "record_id",
    "warning_code",
    "warning_message",
    "field_name",
    "severity",
    "evidence_snippet",
    "schema_type",
]

ABSTRACT_SUMMARY_COLUMNS = [
    "record_id",
    "source_file",
    "schema_type",
    "title",
    "doi",
    "journal",
    "publication_year",
    "article_type",
    "authors",
    "affiliations",
    "keywords",
    "abstract_section_count",
    "structured_abstract",
    "parse_status",
    "parse_warnings",
    "llm_trace_flag",
    "tortured_phrase_flag",
    "nonsense_candidate_flag",
    "numerical_contradiction_flag",
    "design_contradiction_flag",
    "unverifiable_trial_flag",
    "template_flag",
    "template_confidence",
    "template_review_priority",
    "matched_abstract_count",
    "strongest_matched_record_id",
    "strongest_matched_source_file",
    "strongest_matched_title",
    "strongest_match_pair_id",
    "strongest_match_supporting_types",
    "strongest_match_sections",
    "strongest_match_sentence_count",
    "strongest_match_shared_text_coverage",
    "strongest_match_original_text_similarity",
    "strongest_match_masked_skeleton_similarity",
    "strongest_match_ngram_similarity",
    "strongest_match_high_value_section_similarity",
    "strongest_match_weighted_section_similarity",
    "strongest_match_variable_substitutions",
    "strongest_match_relationship_context",
    "strongest_match_evidence_excerpt",
    "strongest_match_review_status",
    "primary_template_pattern",
    "matched_sections",
    "template_cluster_flag",
    "template_family_id",
    "template_family_size",
    "template_family_confidence",
    "template_evidence_summary",
    "llm_trace_count",
    "tortured_phrase_count",
    "nonsense_candidate_count",
    "numerical_contradiction_count",
    "design_contradiction_count",
    "unverifiable_trial_count",
    "total_finding_count",
    "highest_severity",
    "overall_content_risk",
    "review_required",
    "review_reason",
]


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, list):
        if not value:
            return ""
        if all(isinstance(item, dict) for item in value):
            if all({"section", "text"}.issubset(item.keys()) for item in value):
                parts = []
                for item in value:
                    section = normalize_whitespace(item.get("section", ""))
                    text = normalize_whitespace(item.get("text", ""))
                    if section and text:
                        parts.append(f"{section}: {text}")
                    elif text:
                        parts.append(text)
                if parts:
                    return " || ".join(parts)
            return json.dumps(value, ensure_ascii=False, sort_keys=True)
        return " | ".join(normalize_whitespace(str(item)) for item in value if normalize_whitespace(str(item)))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return normalize_whitespace(str(value))


def _write_table(
    ws,
    rows: Sequence[dict[str, Any]],
    columns: Sequence[str],
    start_row: int = 1,
    start_col: int = 1,
    auto_filter: bool = True,
) -> tuple[int, int]:
    row_idx = start_row
    col_idx = start_col
    for offset, column in enumerate(columns, start=col_idx):
        cell = ws.cell(row=row_idx, column=offset, value=column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in rows:
        row_idx += 1
        for offset, column in enumerate(columns, start=col_idx):
            value = _stringify(row.get(column, ""))
            cell = ws.cell(row=row_idx, column=offset, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = THIN_FONT
    if auto_filter and rows:
        ws.auto_filter.ref = f"{ws.cell(start_row, start_col).coordinate}:{ws.cell(row=row_idx, column=col_idx + len(columns) - 1).coordinate}"
    return row_idx, col_idx + len(columns) - 1


def _write_key_value_block(ws, items: Sequence[tuple[str, Any]], start_row: int = 1, title: str | None = None) -> int:
    row = start_row
    if title:
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True)
        row += 1
    for key, value in items:
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row, column=2, value=_stringify(value))
        row += 1
    return row


def _auto_size_columns(ws, max_width: int = 60) -> None:
    widths: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value)
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), min(len(text), max_width))
    for column_letter, width in widths.items():
        ws.column_dimensions[column_letter].width = max(12, min(width + 2, max_width))


def _dashboard_block(ws, title: str, rows: list[dict[str, Any]], columns: list[str], start_row: int) -> int:
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
    end_row, _ = _write_table(ws, rows, columns, start_row=start_row + 1, auto_filter=False)
    return end_row + 2


def _dashboard_section(value: Any) -> str:
    section = normalize_whitespace(str(value)).lower()
    if section == "title":
        return "Title"
    for label in ("background", "methods", "results", "conclusions"):
        if label.rstrip("s") in section:
            return label.title()
    if section == "cross_document":
        return "Cross-document"
    return "Abstract"


def _write_dashboard(
    workbook: Workbook,
    abstract_summary_rows: list[dict[str, Any]],
    findings_rows: list[dict[str, Any]],
    cluster_rows: list[dict[str, Any]],
    parse_warning_rows: list[dict[str, Any]],
) -> None:
    ws = workbook.create_sheet("Dashboard")
    row = 1

    priority_counts = Counter(row["overall_content_risk"] for row in abstract_summary_rows)
    row = _dashboard_block(
        ws,
        "Abstracts by review priority",
        [{"review_priority": priority, "count": priority_counts[priority]} for priority in ("High", "Medium", "Low", "None")],
        ["review_priority", "count"],
        row,
    )

    check_names = {
        "llm_response_trace": "llm_trace",
        "tortured_phrase": "tortured_phrase",
        "template_cluster": "template",
        "nonsense_candidate": "nonsense_candidate",
    }
    check_counts = Counter(item.get("check_type") or check_names.get(item.get("detector_type", ""), item.get("detector_type", "")) for item in findings_rows)
    row = _dashboard_block(
        ws,
        "Findings by check type",
        [{"check_type": check_type, "count": count} for check_type, count in sorted(check_counts.items())],
        ["check_type", "count"],
        row,
    )

    clusters: dict[str, tuple[int, str]] = {}
    for item in cluster_rows:
        cluster_id = str(item.get("template_family_id", ""))
        if not cluster_id:
            continue
        size = int(item.get("member_count") or 0)
        representative = str(item.get("representative_record_id", ""))
        current = clusters.get(cluster_id)
        clusters[cluster_id] = (max(size, current[0] if current else 0), min(representative, current[1]) if current else representative)
    size_counts = Counter("5+" if size >= 5 else str(size) for size, _ in clusters.values())
    row = _dashboard_block(
        ws,
        "Template cluster summary",
        [
            {"metric": "total_clusters", "count": len(clusters)},
            *({"metric": f"size_{bucket}", "count": size_counts[bucket]} for bucket in ("2", "3", "4", "5+")),
        ],
        ["metric", "count"],
        row,
    )
    row = _dashboard_block(
        ws,
        "Largest 10 template clusters",
        [
            {"template_cluster_id": cluster_id, "cluster_size": size, "representative_record_id": representative}
            for cluster_id, (size, representative) in sorted(clusters.items(), key=lambda item: (-item[1][0], item[0]))[:10]
        ],
        ["template_cluster_id", "cluster_size", "representative_record_id"],
        row,
    )

    warning_counts = Counter(str(item.get("warning_code", "")) for item in parse_warning_rows)
    row = _dashboard_block(
        ws,
        "Parse failures and warnings by type",
        [{"warning_type": warning_type, "count": count} for warning_type, count in sorted(warning_counts.items())]
        or [{"warning_type": "NONE", "count": 0}],
        ["warning_type", "count"],
        row,
    )

    section_counts = Counter(_dashboard_section(item.get("section_or_field", "")) for item in findings_rows)
    _dashboard_block(
        ws,
        "Findings by abstract section",
        [{"section": section, "count": count} for section, count in sorted(section_counts.items())],
        ["section", "count"],
        row,
    )
    ws.freeze_panes = "A3"
    _auto_size_columns(ws)


def write_jsonl(path: str | Path, rows: Iterable[dict[str, Any]]) -> Path:
    resolved = ensure_parent_dir(path)
    with resolved.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True))
            handle.write("\n")
    return resolved


def write_csv(path: str | Path, rows: Sequence[dict[str, Any]], columns: Sequence[str]) -> Path:
    resolved = ensure_parent_dir(path)
    with resolved.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(columns))
        writer.writeheader()
        for row in rows:
            writer.writerow({column: _stringify(row.get(column, "")) for column in columns})
    return resolved


def write_workbook(
    path: str | Path,
    *,
    inventory_rows: list[dict[str, Any]],
    root_summary_rows: list[dict[str, Any]],
    abstract_summary_rows: list[dict[str, Any]],
    findings_rows: list[dict[str, Any]],
    pair_rows: list[dict[str, Any]],
    cluster_rows: list[dict[str, Any]],
    dictionary_rows: list[dict[str, Any]],
    parse_warning_rows: list[dict[str, Any]],
    run_metadata_rows: list[tuple[str, Any]],
) -> Path:
    resolved = ensure_parent_dir(path)
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    _write_dashboard(workbook, abstract_summary_rows, findings_rows, cluster_rows, parse_warning_rows)

    # Data Inventory sheet
    ws = workbook.create_sheet("Data Inventory")
    metrics = [item for item in root_summary_rows if not item[0].startswith("root_element_")]
    root_elements = [item for item in root_summary_rows if item[0].startswith("root_element_")]
    ws.cell(row=1, column=1, value="metric").fill = HEADER_FILL
    ws.cell(row=1, column=1).font = HEADER_FONT
    ws.cell(row=1, column=2, value="value").fill = HEADER_FILL
    ws.cell(row=1, column=2).font = HEADER_FONT
    row = 2
    for metric, value in metrics:
        ws.cell(row=row, column=1, value=metric).font = Font(bold=True)
        ws.cell(row=row, column=2, value=_stringify(value))
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="root_element").fill = HEADER_FILL
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=2, value="count").fill = HEADER_FILL
    ws.cell(row=row, column=2).font = HEADER_FONT
    row += 1
    for root_metric, value in root_elements:
        ws.cell(row=row, column=1, value=root_metric.replace("root_element_", "")).font = THIN_FONT
        ws.cell(row=row, column=2, value=value).font = THIN_FONT
        row += 1
    row += 1
    field_columns = [
        "field_name",
        "primary_xml_path",
        "present_count",
        "present_pct",
        "example_value",
        "useful_for_poc",
        "notes",
    ]
    _write_table(ws, inventory_rows, field_columns, start_row=row)
    ws.freeze_panes = f"A{row+1}"
    _auto_size_columns(ws)

    # Abstract Summary
    ws = workbook.create_sheet("Abstract Summary")
    _write_table(ws, abstract_summary_rows, ABSTRACT_SUMMARY_COLUMNS, start_row=1)
    ws.freeze_panes = "A2"
    _auto_size_columns(ws)

    # Integrity Findings
    ws = workbook.create_sheet("Integrity Findings")
    _write_table(
        ws,
        findings_rows
        or [
            {
                "finding_id": "",
                "record_id": "",
                "source_file": "",
                "detector_type": "",
                "category": "",
                "matched_text": "",
                "expected_term": "",
                "evidence_snippet": "No integrity findings detected in this run.",
                "section_or_field": "",
                "severity": "",
                "confidence": "",
                "validation_status": "",
                "validation_reason": "",
                "validated_by": "",
                "rule_id": "",
                "template_cluster_id": "",
                "cluster_size": "",
                "similar_record_ids": "",
                "similarity_score": "",
                "cluster_severity": "",
                "shared_skeleton_excerpt": "No integrity findings detected in this run.",
                "metadata_context": "",
            }
        ],
        FINDINGS_COLUMNS,
        start_row=1,
    )
    ws.freeze_panes = "A2"
    _auto_size_columns(ws)

    # Template Pair Evidence
    ws = workbook.create_sheet("Template Pairs")
    _write_table(ws, pair_rows, PAIR_COLUMNS, start_row=1)
    ws.freeze_panes = "A2"
    _auto_size_columns(ws)

    # Template Clusters
    ws = workbook.create_sheet("Template Clusters")
    _write_table(ws, cluster_rows, FAMILY_COLUMNS, start_row=1)
    ws.freeze_panes = "A2"
    _auto_size_columns(ws)

    # Pattern Dictionary
    ws = workbook.create_sheet("Pattern Dictionary")
    _write_table(ws, dictionary_rows, DICTIONARY_COLUMNS, start_row=1)
    ws.freeze_panes = "A2"
    _auto_size_columns(ws)

    # Parse Warnings
    ws = workbook.create_sheet("Parse Warnings")
    _write_table(ws, parse_warning_rows, WARNING_COLUMNS, start_row=1)
    ws.freeze_panes = "A2"
    _auto_size_columns(ws)

    # Run Metadata
    ws = workbook.create_sheet("Run Metadata")
    _write_key_value_block(ws, run_metadata_rows, start_row=1)
    _auto_size_columns(ws)

    workbook.save(resolved)
    return resolved
